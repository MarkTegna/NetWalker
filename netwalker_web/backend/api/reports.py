"""
Report generation API endpoints

Author: Mark Oldham
"""

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import FileResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import config

router = APIRouter()


def generate_device_report(devices: list, output_file: str):
    """Generate device inventory Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device Inventory"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ["Device Name", "Platform", "Hardware Model", "Serial Number", 
               "Software Version", "IP Address", "Capabilities", "Status"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    for row_num, device in enumerate(devices, 2):
        ws.cell(row=row_num, column=1, value=device.get('device_name'))
        ws.cell(row=row_num, column=2, value=device.get('platform'))
        ws.cell(row=row_num, column=3, value=device.get('hardware_model'))
        ws.cell(row=row_num, column=4, value=device.get('serial_number'))
        ws.cell(row=row_num, column=5, value=device.get('software_version'))
        ws.cell(row=row_num, column=6, value=device.get('ip_address'))
        ws.cell(row=row_num, column=7, value=device.get('capabilities'))
        ws.cell(row=row_num, column=8, value=device.get('status'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)


@router.get("/reports/devices")
async def generate_devices_report(
    request: Request,
    device_name: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    hardware_model: Optional[str] = Query(None),
    serial_number: Optional[str] = Query(None),
    capabilities: Optional[str] = Query(None),
    ip_address: Optional[str] = Query(None),
    software_version: Optional[str] = Query(None)
):
    """
    Generate device inventory Excel report with optional filters
    
    Returns:
        Excel file download
    """
    try:
        # Build filters dictionary
        filters = {}
        if device_name:
            filters['device_name'] = device_name
        if platform:
            filters['platform'] = platform
        if hardware_model:
            filters['hardware_model'] = hardware_model
        if serial_number:
            filters['serial_number'] = serial_number
        if capabilities:
            filters['capabilities'] = capabilities
        if ip_address:
            filters['ip_address'] = ip_address
        if software_version:
            filters['software_version'] = software_version
        
        # Get devices with filters
        devices = request.app.state.device_queries.get_all_devices(
            limit=10000,
            filters=filters if filters else None
        )
        
        # Generate filename with YYYYMMDD-HH-MM format
        timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
        filename = f"Device_Inventory_{timestamp}.xlsx"
        output_file = config.REPORTS_DIR / filename
        
        # Generate report
        generate_device_report(devices, str(output_file))
        
        return FileResponse(
            path=output_file,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/stacks")
async def generate_stacks_report(request: Request):
    """
    Generate stack members Excel report
    
    Returns:
        Excel file download
    """
    try:
        # Get all stacks
        stacks = request.app.state.stack_queries.get_all_stacks()
        
        # Generate filename with YYYYMMDD-HH-MM format
        timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
        filename = f"Stack_Members_{timestamp}.xlsx"
        output_file = config.REPORTS_DIR / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stack Members"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Headers
        headers = ["Device Name", "Switch Number", "Role", "Priority", 
                   "Hardware Model", "Serial Number", "MAC Address", "State"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row_num = 2
        for stack in stacks:
            members = request.app.state.stack_queries.get_stack_members(stack['device_id'])
            for member in members:
                ws.cell(row=row_num, column=1, value=stack['device_name'])
                ws.cell(row=row_num, column=2, value=member.get('switch_number'))
                ws.cell(row=row_num, column=3, value=member.get('role'))
                ws.cell(row=row_num, column=4, value=member.get('priority'))
                ws.cell(row=row_num, column=5, value=member.get('hardware_model'))
                ws.cell(row=row_num, column=6, value=member.get('serial_number'))
                ws.cell(row=row_num, column=7, value=member.get('mac_address'))
                ws.cell(row=row_num, column=8, value=member.get('state'))
                row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_file)
        
        return FileResponse(
            path=output_file,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
