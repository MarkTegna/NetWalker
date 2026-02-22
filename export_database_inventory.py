"""
Export all devices from NetWalker database to Excel
Author: Mark Oldham
"""

import pyodbc
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def decode_password(encoded_password):
    """Decode the encrypted password"""
    if encoded_password.startswith("ENC:"):
        encoded_password = encoded_password[4:]
    return base64.b64decode(encoded_password).decode('utf-8')

def connect_to_database():
    """Connect to the NetWalker database"""
    # Database configuration
    server = "eit-prisqldb01.tgna.tegna.com"
    port = 1433
    database = "NetWalker"
    username = "NetWalker"
    encoded_password = "ENC:Rmx1ZmZ5QnVubnlIaXRieWFCdXM="
    
    # Decode password
    password = decode_password(encoded_password)
    
    # Build connection string
    conn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={server},{port};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        f"TrustServerCertificate=yes;"
    )
    
    try:
        connection = pyodbc.connect(conn_str, timeout=30)
        print(f"Connected to database: {database}")
        return connection
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

def export_devices_to_excel():
    """Export all devices from database to Excel"""
    
    # Connect to database
    conn = connect_to_database()
    if not conn:
        print("Failed to connect to database")
        return
    
    try:
        # Query to get all devices with their details
        query = """
        SELECT 
            d.device_name,
            di.ip_address,
            d.platform,
            dv.software_version,
            d.hardware_model,
            d.serial_number,
            d.capabilities,
            d.first_seen,
            d.last_seen,
            d.status
        FROM devices d
        LEFT JOIN (
            SELECT device_id, ip_address,
                   ROW_NUMBER() OVER (PARTITION BY device_id ORDER BY last_seen DESC) as rn
            FROM device_interfaces
        ) di ON d.device_id = di.device_id AND di.rn = 1
        LEFT JOIN (
            SELECT device_id, software_version,
                   ROW_NUMBER() OVER (PARTITION BY device_id ORDER BY last_seen DESC) as rn
            FROM device_versions
        ) dv ON d.device_id = dv.device_id AND dv.rn = 1
        ORDER BY d.device_name
        """
        
        print("Querying database for all devices...")
        
        cursor = conn.cursor()
        cursor.execute(query)
        
        # Get column names
        columns = [column[0] for column in cursor.description]
        
        # Fetch all rows
        rows = cursor.fetchall()
        
        print(f"Found {len(rows)} devices in database")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Device Inventory"
        
        # Write header row with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook with date-coded filename
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
        output_file = f"Device_Inventory_{timestamp}.xlsx"
        wb.save(output_file)
        
        print(f"Successfully exported {len(rows)} devices to {output_file}")
        
        # Print summary statistics
        print("\nSummary:")
        print(f"  Total devices: {len(rows)}")
        
        # Count unique sites and platforms
        platforms = set()
        statuses = set()
        for row in rows:
            if row[2]:  # platform column
                platforms.add(row[2])
            if row[9]:  # status column
                statuses.add(row[9])
        
        print(f"  Unique platforms: {len(platforms)}")
        print(f"  Device statuses: {', '.join(sorted(statuses))}")
        
    except Exception as e:
        print(f"Error exporting devices: {e}")
        import traceback
        traceback.print_exc()
    finally:
        conn.close()
        print("Database connection closed")

if __name__ == "__main__":
    print("NetWalker Database Device Export")
    print("=" * 50)
    export_devices_to_excel()

if __name__ == "__main__":
    print("NetWalker Database Device Export")
    print("=" * 50)
    export_devices_to_excel()
