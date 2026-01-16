import openpyxl

wb = openpyxl.load_workbook('dist/NetWalker_v0.4.4/reports/Inventory_20260115-13-52.xlsx')
ws = wb['Device Inventory']

print('Headers:')
for i, cell in enumerate(ws[1]):
    print(f'  {i}: {cell.value}')

print('\nRow 2 (BORO-UW01):')
for i, cell in enumerate(ws[2]):
    print(f'  {i}: {cell.value}')
