"""
Export devices with 2 or more connection failures to Excel spreadsheet
"""
import sys
sys.path.insert(0, '.')

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from netwalker.config.config_manager import ConfigurationManager
from netwalker.database.database_manager import DatabaseManager

# Load config
config_manager = ConfigurationManager('netwalker.ini')
full_config = config_manager.load_configuration()
db_config = full_config.get('database')

# Connect to database
db = DatabaseManager(db_config)
if not db.connect():
    print("Failed to connect to database")
    sys.exit(1)

try:
    cursor = db.connection.cursor()
    
    # Query devices with 2 or more connection failures
    cursor.execute("""
        SELECT 
            d.device_id,
            d.device_name,
            d.serial_number,
            d.platform,
            d.hardware_model,
            d.capabilities,
            d.connection_failures,
            d.first_seen,
            d.last_seen,
            d.status,
            i.ip_address,
            i.interface_type
        FROM devices d
        LEFT JOIN (
            SELECT device_id, ip_address, interface_type,
                   ROW_NUMBER() OVER (PARTITION BY device_id ORDER BY 
                       CASE WHEN interface_type = 'Management' THEN 0 ELSE 1 END,
                       interface_id) as rn
            FROM device_interfaces
        ) i ON d.device_id = i.device_id AND i.rn = 1
        WHERE d.connection_failures >= 2
        ORDER BY d.connection_failures DESC, d.device_name
    """)
    
    rows = cursor.fetchall()
    
    if not rows:
        print("No devices found with 2 or more connection failures")
        cursor.close()
        db.disconnect()
        sys.exit(0)
    
    print(f"Found {len(rows)} device(s) with 2 or more connection failures")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Failed Connections"
    
    # Define headers
    headers = [
        "Device ID",
        "Device Name",
        "IP Address",
        "Serial Number",
        "Platform",
        "Hardware Model",
        "Capabilities",
        "Connection Failures",
        "First Seen",
        "Last Seen",
        "Status",
        "Interface Type"
    ]
    
    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        device_id = row_data[0]
        device_name = row_data[1]
        serial_number = row_data[2]
        platform = row_data[3]
        hardware_model = row_data[4]
        capabilities = row_data[5]
        connection_failures = row_data[6]
        first_seen = row_data[7]
        last_seen = row_data[8]
        status = row_data[9]
        ip_address = row_data[10] if row_data[10] else "N/A"
        interface_type = row_data[11] if row_data[11] else "N/A"
        
        # Write row data
        ws.cell(row=row_num, column=1).value = device_id
        ws.cell(row=row_num, column=2).value = device_name
        ws.cell(row=row_num, column=3).value = ip_address
        ws.cell(row=row_num, column=4).value = serial_number
        ws.cell(row=row_num, column=5).value = platform
        ws.cell(row=row_num, column=6).value = hardware_model
        ws.cell(row=row_num, column=7).value = capabilities
        ws.cell(row=row_num, column=8).value = connection_failures
        ws.cell(row=row_num, column=9).value = first_seen
        ws.cell(row=row_num, column=10).value = last_seen
        ws.cell(row=row_num, column=11).value = status
        ws.cell(row=row_num, column=12).value = interface_type
        
        # Highlight rows with high failure counts
        if connection_failures >= 5:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        for row_num in range(2, len(rows) + 2):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[chr(64 + col_num)].width = min(max_length + 2, 50)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
    filename = f"Failed_Connections_{timestamp}.xlsx"
    
    # Save workbook
    wb.save(filename)
    print(f"\nSpreadsheet created successfully: {filename}")
    print(f"Total devices: {len(rows)}")
    
    # Print summary statistics
    high_failure_count = sum(1 for row in rows if row[6] >= 5)
    if high_failure_count > 0:
        print(f"Devices with 5+ failures (highlighted in red): {high_failure_count}")
    
    cursor.close()
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
finally:
    db.disconnect()
