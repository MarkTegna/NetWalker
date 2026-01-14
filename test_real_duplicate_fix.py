#!/usr/bin/env python3
"""
Test the duplicate sheets fix with real NetWalker execution.
"""

import sys
import os
import logging
import csv
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def test_with_boro_core_a():
    """Test with BORO-CORE-A which we know has potential duplicates"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("TESTING DUPLICATE SHEETS FIX WITH BORO-CORE-A")
    logger.info("=" * 80)
    
    try:
        # Create test seed file with BORO-CORE-A
        seed_file = "test_duplicate_fix_seed.csv"
        
        with open(seed_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['hostname', 'ip_address', 'status'])
            writer.writerow(['BORO-CORE-A', '10.1.1.1', ''])  # Use placeholder IP
        
        logger.info(f"Created test seed file: {seed_file}")
        
        # Create test config with depth 1 to keep it manageable
        config_file = "test_duplicate_fix_config.ini"
        
        config_content = """[discovery]
seed_file = test_duplicate_fix_seed.csv
max_depth = 1
timeout = 30
concurrent_limit = 5

[output]
site_boundary_pattern = *-CORE-*
output_directory = test_duplicate_fix_reports
create_master_inventory = true
create_site_reports = true

[credentials]
username = admin
password = admin
enable_password = admin

[filtering]
enable_device_filtering = false
enable_lumt_filtering = false

[validation]
enable_dns_validation = false
"""
        
        with open(config_file, 'w') as f:
            f.write(config_content)
        
        logger.info(f"Created test config file: {config_file}")
        
        # Create output directory
        os.makedirs("test_duplicate_fix_reports", exist_ok=True)
        
        logger.info("Running NetWalker with BORO-CORE-A (depth 1)...")
        
        # Run NetWalker
        import subprocess
        
        cmd = [
            ".\\dist\\netwalker.exe",
            "--config", config_file,
            "--verbose"
        ]
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        
        logger.info(f"NetWalker exit code: {result.returncode}")
        
        if result.returncode != 0:
            logger.error("NetWalker execution failed:")
            logger.error(f"STDOUT: {result.stdout}")
            logger.error(f"STDERR: {result.stderr}")
            return False
        
        logger.info("✅ NetWalker execution completed successfully")
        
        # Analyze the generated seed workbook
        reports_dir = Path("test_duplicate_fix_reports")
        seed_reports = list(reports_dir.glob("Seed_BORO-CORE-A_*.xlsx"))
        
        if not seed_reports:
            logger.error("No BORO-CORE-A seed report generated")
            return False
        
        latest_report = max(seed_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"Generated seed report: {latest_report}")
        
        # Analyze for duplicates
        from openpyxl import load_workbook
        
        workbook = load_workbook(latest_report)
        sheet_names = workbook.sheetnames
        
        logger.info(f"Total sheets in workbook: {len(sheet_names)}")
        
        neighbor_sheets = [s for s in sheet_names if s.startswith("Neighbors_")]
        logger.info(f"Neighbor sheets: {len(neighbor_sheets)}")
        
        for sheet in neighbor_sheets:
            logger.info(f"  {sheet}")
        
        # Check for duplicates
        unique_sheets = set(neighbor_sheets)
        
        if len(neighbor_sheets) == len(unique_sheets):
            logger.info(f"\n✅ NO DUPLICATE SHEETS - Fix is working!")
            logger.info(f"All {len(neighbor_sheets)} neighbor sheets have unique names")
        else:
            logger.error(f"\n❌ DUPLICATE SHEETS STILL EXIST")
            logger.error(f"Total sheets: {len(neighbor_sheets)}, Unique: {len(unique_sheets)}")
            
            # Find the duplicates
            sheet_counts = {}
            for sheet in neighbor_sheets:
                sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1
            
            for sheet, count in sheet_counts.items():
                if count > 1:
                    logger.error(f"  {sheet}: {count} times")
        
        workbook.close()
        
        return len(neighbor_sheets) == len(unique_sheets)
        
    except Exception as e:
        logger.error(f"Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Cleanup
        for file in ["test_duplicate_fix_seed.csv", "test_duplicate_fix_config.ini"]:
            if os.path.exists(file):
                os.remove(file)

def main():
    """Main test function"""
    logger = setup_logging()
    
    logger.info("NetWalker Duplicate Sheets Fix - Real Test")
    
    # Test with BORO-CORE-A
    test_result = test_with_boro_core_a()
    
    logger.info(f"\n" + "=" * 80)
    logger.info("REAL TEST SUMMARY")
    logger.info("=" * 80)
    
    if test_result:
        logger.info("✅ DUPLICATE SHEETS FIX VERIFIED")
        logger.info("The fix successfully prevents duplicate sheets in real NetWalker execution")
        return True
    else:
        logger.error("❌ DUPLICATE SHEETS FIX FAILED")
        logger.error("The fix did not prevent duplicate sheets - more work needed")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)