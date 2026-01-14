#!/usr/bin/env python3
"""
Analyze the most recent discovery to see what was discovered.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def analyze_inventory(inventory_path):
    """Analyze an inventory file"""
    logger = logging.getLogger(__name__)
    
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(inventory_path)
        
        # Find the device inventory sheet
        inventory_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'inventory' in sheet_name.lower() or 'device' in sheet_name.lower():
                inventory_sheet = workbook[sheet_name]
                break
        
        if not inventory_sheet:
            logger.error("Could not find inventory sheet")
            return None
        
        logger.info(f"Analyzing sheet: {inventory_sheet.title}")
        
        # Count devices (skip header row)
        device_count = 0
        devices = []
        
        for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # If first column has data
                device_count += 1
                # Assuming columns are: Key, Hostname, IP, Platform, Status, etc.
                device_info = {
                    'key': row[0] if len(row) > 0 else '',
                    'hostname': row[1] if len(row) > 1 else '',
                    'ip': row[2] if len(row) > 2 else '',
                    'platform': row[3] if len(row) > 3 else '',
                    'status': row[8] if len(row) > 8 else ''
                }
                devices.append(device_info)
        
        workbook.close()
        
        return {
            'count': device_count,
            'devices': devices
        }
        
    except Exception as e:
        logger.error(f"Error analyzing inventory: {e}")
        import traceback
        traceback.print_exc()
        return None

def main():
    """Main analysis function"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("ANALYZING RECENT DISCOVERY")
    logger.info("=" * 80)
    
    # Find the most recent inventory
    reports_dir = Path("dist/NetWalker_v0.3.29/reports")
    inventory_files = list(reports_dir.glob("Inventory_*.xlsx"))
    
    if not inventory_files:
        logger.error("No inventory files found")
        return False
    
    latest_inventory = max(inventory_files, key=lambda p: p.stat().st_mtime)
    logger.info(f"Latest inventory: {latest_inventory.name}")
    
    result = analyze_inventory(latest_inventory)
    
    if result:
        logger.info(f"\n{'=' * 80}")
        logger.info(f"DISCOVERY RESULTS")
        logger.info(f"{'=' * 80}")
        logger.info(f"Total devices discovered: {result['count']}")
        
        if result['count'] > 0:
            logger.info(f"\nDevices:")
            for i, device in enumerate(result['devices'][:20], 1):  # Show first 20
                logger.info(f"  {i}. {device['hostname']} ({device['ip']}) - {device['platform']} - Status: {device['status']}")
            
            if result['count'] > 20:
                logger.info(f"  ... and {result['count'] - 20} more devices")
        
        # Check if WTSP-CORE-A was discovered
        wtsp_found = any('WTSP-CORE-A' in device['hostname'].upper() for device in result['devices'])
        
        logger.info(f"\n{'=' * 80}")
        logger.info(f"WTSP-CORE-A STATUS")
        logger.info(f"{'=' * 80}")
        
        if wtsp_found:
            logger.info("✅ WTSP-CORE-A was discovered")
            wtsp_device = next(d for d in result['devices'] if 'WTSP-CORE-A' in d['hostname'].upper())
            logger.info(f"  Hostname: {wtsp_device['hostname']}")
            logger.info(f"  IP: {wtsp_device['ip']}")
            logger.info(f"  Platform: {wtsp_device['platform']}")
            logger.info(f"  Status: {wtsp_device['status']}")
        else:
            logger.error("❌ WTSP-CORE-A was NOT discovered")
        
        return True
    
    return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)