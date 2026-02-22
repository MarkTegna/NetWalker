"""
Quick script to check if Stack Members sheet has data in the Excel file
"""
import openpyxl
import sys

if len(sys.argv) < 2:
    print("Usage: python check_excel_stack_members.py <excel_file>")
    sys.exit(1)

excel_file = sys.argv[1]

try:
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    
    print(f"\nExcel file: {excel_file}")
    print(f"Available sheets: {wb.sheetnames}")
    
    if 'Stack Members' in wb.sheetnames:
        ws = wb['Stack Members']
        print(f"\nStack Members sheet found!")
        print(f"Max row: {ws.max_row}")
        print(f"Max column: {ws.max_column}")
        
        # Read first 10 rows
        print("\nFirst 10 rows:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row_idx}: {' | '.join(row_data)}")
    else:
        print("\nStack Members sheet NOT found!")
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
