import openpyxl

# Load the most recent discovery report
wb = openpyxl.load_workbook('reports/Discovery_20260113-17-57.xlsx')
ws_vlan = wb['VLAN Inventory']

print("=" * 80)
print("LATEST DISCOVERY - BORO-MDF Devices VLAN Data")
print("=" * 80)

# Group VLANs by device
device_vlans = {}
for row in range(2, ws_vlan.max_row + 1):
    device_name = ws_vlan.cell(row, 1).value
    if device_name and 'BORO-MDF' in str(device_name):
        if device_name not in device_vlans:
            device_vlans[device_name] = {'vlans': [], 'status': None, 'timestamp': None}
        vlan_id = ws_vlan.cell(row, 2).value
        vlan_name = ws_vlan.cell(row, 3).value
        port_count = ws_vlan.cell(row, 4).value
        status = ws_vlan.cell(row, 9).value
        timestamp = ws_vlan.cell(row, 8).value
        device_vlans[device_name]['vlans'].append((vlan_id, vlan_name, port_count))
        device_vlans[device_name]['status'] = status
        device_vlans[device_name]['timestamp'] = timestamp

# Check each device
for device in sorted(device_vlans.keys()):
    print(f"\n{device}:")
    print(f"  Collection Status: {device_vlans[device]['status']}")
    print(f"  Collection Time: {device_vlans[device]['timestamp']}")
    print(f"  Total VLANs: {len(device_vlans[device]['vlans'])}")
    
    has_216 = False
    has_461 = False
    
    for vlan_id, vlan_name, port_count in sorted(device_vlans[device]['vlans']):
        if vlan_id == 216:
            has_216 = True
            print(f"  ✅ VLAN 216: {vlan_name} (Ports: {port_count})")
        elif vlan_id == 461:
            has_461 = True
            print(f"  ✅ VLAN 461: {vlan_name} (Ports: {port_count})")
    
    if not has_216:
        print(f"  ❌ VLAN 216: NOT FOUND")
    if not has_461:
        print(f"  ❌ VLAN 461: NOT FOUND")

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
for device in sorted(device_vlans.keys()):
    has_216 = any(v[0] == 216 for v in device_vlans[device]['vlans'])
    has_461 = any(v[0] == 461 for v in device_vlans[device]['vlans'])
    print(f"{device}: VLAN 216={'✅' if has_216 else '❌'}, VLAN 461={'✅' if has_461 else '❌'}")
