"""
Extract Device Names from All_Equipment-20250815-200140.xlsx
Compares to database and writes new devices to seed_file.csv

Author: Mark Oldham
"""

import sys
import logging
import csv
import base64
from pathlib import Path
from typing import Set, List
import pyodbc
import configparser
import openpyxl

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class AllEquipmentScanner:
    """Extracts device names from All_Equipment Excel file"""
    
    def __init__(self, config_file: str = "netwalker.ini"):
        """Initialize scanner with database config"""
        self.config = self._load_config(config_file)
        self.connection = None
    
    def _load_config(self, config_file: str) -> dict:
        """Load database configuration from INI file"""
        config = configparser.ConfigParser()
        
        if not Path(config_file).exists():
            logger.error(f"Config file not found: {config_file}")
            sys.exit(1)
        
        config.read(config_file)
        
        # Extract database config
        db_config = {
            'server': config.get('database', 'server'),
            'port': config.get('database', 'port'),
            'database': config.get('database', 'database'),
            'username': config.get('database', 'username'),
            'password': config.get('database', 'password'),
        }
        
        # Decrypt password if encrypted
        if db_config['password'].startswith('ENC:'):
            encoded = db_config['password'][4:]
            db_config['password'] = base64.b64decode(encoded.encode()).decode()
        
        return db_config
    
    def connect_database(self) -> bool:
        """Connect to NetWalker database"""
        try:
            connection_string = (
                f"DRIVER={{SQL Server}};"
                f"SERVER={self.config['server']},{self.config['port']};"
                f"DATABASE={self.config['database']};"
                f"UID={self.config['username']};"
                f"PWD={self.config['password']};"
                "TrustServerCertificate=yes;"
                "Connection Timeout=30;"
            )
            
            self.connection = pyodbc.connect(connection_string)
            logger.info(f"Connected to database: {self.config['server']}/{self.config['database']}")
            return True
            
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            return False
    
    def disconnect_database(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            logger.info("Database connection closed")
    
    def get_existing_devices(self) -> Set[str]:
        """Get all device names from database"""
        try:
            cursor = self.connection.cursor()
            
            query = "SELECT DISTINCT device_name FROM devices"
            cursor.execute(query)
            
            devices = {row[0].upper() for row in cursor.fetchall()}
            cursor.close()
            
            logger.info(f"Retrieved {len(devices)} existing devices from database")
            return devices
            
        except Exception as e:
            logger.error(f"Failed to retrieve devices: {e}")
            return set()
    
    def extract_devices_from_excel(self, file_path: str) -> List[str]:
        """
        Extract device names from Excel file
        
        Args:
            file_path: Path to All_Equipment Excel file
            
        Returns:
            List of device names
        """
        device_names = []
        
        try:
            logger.info(f"Reading Excel file: {file_path}")
            
            # Open workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            logger.info(f"Worksheet: {ws.title}")
            logger.info(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
            
            # Find hostname column (should be column 1 based on inspection)
            header_row = 1
            hostname_col = None
            
            # Check header row
            for col in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(header_row, col).value
                if cell_value:
                    cell_value_upper = str(cell_value).strip().upper()
                    logger.debug(f"Header col {col}: {cell_value}")
                    
                    if 'HOSTNAME' in cell_value_upper or cell_value_upper == 'HOSTNAME':
                        hostname_col = col
                        logger.info(f"Found hostname column at index {col}: {cell_value}")
                        break
            
            if hostname_col is None:
                logger.error("Could not find hostname column")
                return device_names
            
            # Extract device names from data rows
            for row in range(header_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row, hostname_col).value
                
                if cell_value:
                    device_name = str(cell_value).strip().upper()
                    if device_name and device_name != '':
                        device_names.append(device_name)
                        logger.debug(f"Row {row}: {device_name}")
            
            logger.info(f"Extracted {len(device_names)} device names from Excel file")
            
            wb.close()
            return device_names
            
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            logger.exception("Full exception details:")
            return device_names
    
    def write_new_devices_to_csv(self, new_devices: List[str], output_file: str = "seed_file.csv"):
        """
        Write new devices to CSV file
        
        Args:
            new_devices: List of device names not in database
            output_file: Output CSV filename
        """
        try:
            # Check if file exists
            file_exists = Path(output_file).exists()
            
            # Read existing devices from CSV if file exists
            existing_csv_devices = set()
            if file_exists:
                with open(output_file, 'r', newline='') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if 'hostname' in row:
                            existing_csv_devices.add(row['hostname'].upper())
                logger.info(f"Found {len(existing_csv_devices)} existing devices in CSV")
            
            # Filter out devices already in CSV
            devices_to_add = [d for d in new_devices if d not in existing_csv_devices]
            
            if not devices_to_add:
                logger.info("All new devices are already in the CSV file")
                return
            
            # Open file in append mode
            with open(output_file, 'a', newline='') as f:
                writer = csv.writer(f)
                
                # Write header if new file
                if not file_exists:
                    writer.writerow(['hostname', 'ip_address', 'status'])
                    logger.info(f"Created new file: {output_file}")
                
                # Write device names
                for device_name in sorted(devices_to_add):
                    writer.writerow([device_name, '', 'pending'])
                
            logger.info(f"Wrote {len(devices_to_add)} new devices to {output_file}")
            
        except Exception as e:
            logger.error(f"Failed to write CSV file: {e}")


def main():
    """Main program entry point"""
    
    # Configuration
    EXCEL_FILE = r"D:\MJODev\NetWalker\prodtest_files\book1.xlsx"
    OUTPUT_FILE = "seed_file.csv"
    
    print("=" * 70)
    print("All Equipment Device Extractor")
    print("=" * 70)
    print()
    
    try:
        # Initialize scanner
        scanner = AllEquipmentScanner()
        
        # Connect to database
        logger.info("Connecting to database...")
        if not scanner.connect_database():
            logger.error("Failed to connect to database")
            return 1
        
        # Get existing devices from database
        logger.info("Retrieving existing devices from database...")
        existing_devices = scanner.get_existing_devices()
        
        # Extract devices from Excel file
        logger.info(f"Extracting devices from Excel file: {EXCEL_FILE}")
        found_devices = scanner.extract_devices_from_excel(EXCEL_FILE)
        
        if not found_devices:
            logger.warning("No devices found in Excel file")
            scanner.disconnect_database()
            return 0
        
        # Find new devices (not in database)
        new_devices = [d for d in found_devices if d not in existing_devices]
        
        # Remove duplicates
        new_devices = list(set(new_devices))
        
        logger.info(f"Found {len(found_devices)} total devices in Excel file")
        logger.info(f"Found {len(set(found_devices))} unique devices in Excel file")
        logger.info(f"Found {len(existing_devices)} existing devices in database")
        logger.info(f"Found {len(new_devices)} NEW devices not in database")
        
        if new_devices:
            # Write new devices to CSV
            logger.info(f"Writing new devices to {OUTPUT_FILE}...")
            scanner.write_new_devices_to_csv(new_devices, OUTPUT_FILE)
            
            print()
            print(f"[OK] Found {len(new_devices)} new devices")
            print(f"[OK] Saved to: {OUTPUT_FILE}")
            print()
            print(f"New devices (showing first 50):")
            for device in sorted(new_devices)[:50]:
                print(f"  - {device}")
            
            if len(new_devices) > 50:
                print(f"  ... and {len(new_devices) - 50} more")
        else:
            print()
            print("[INFO] No new devices found - all devices already in database")
        
        # Disconnect
        scanner.disconnect_database()
        
        print()
        print("Extraction complete!")
        return 0
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        return 130
    
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        logger.exception("Full exception details:")
        return 1


if __name__ == '__main__':
    sys.exit(main())
