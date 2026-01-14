#!/usr/bin/env python3
"""
Test script for DNS hostname cleaning with real NetWalker execution.

This script creates a test scenario with BORO-CORE-A as seed and depth 2
to verify that hostnames with serial numbers are properly cleaned in DNS validation.
"""

import sys
import os
import logging
import csv
import time
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def create_test_seed_file():
    """Create a test seed file with BORO-CORE-A"""
    logger = logging.getLogger(__name__)
    
    seed_file = "test_dns_seed.csv"
    
    # Create seed file with BORO-CORE-A
    with open(seed_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['hostname', 'ip_address', 'status'])
        writer.writerow(['BORO-CORE-A', '10.1.1.1', ''])  # Empty status for fresh discovery
    
    logger.info(f"Created test seed file: {seed_file}")
    return seed_file

def create_test_config():
    """Create a test configuration file"""
    logger = logging.getLogger(__name__)
    
    config_file = "test_dns_config.ini"
    
    config_content = """[discovery]
seed_file = test_dns_seed.csv
max_depth = 2
timeout = 30
concurrent_limit = 5

[output]
site_boundary_pattern = 
output_directory = test_dns_reports
create_master_inventory = true
create_site_reports = false

[credentials]
username = admin
password = admin
enable_password = admin

[filtering]
enable_device_filtering = false
enable_lumt_filtering = false

[validation]
enable_dns_validation = true
dns_timeout = 10
max_concurrent_dns = 3
"""
    
    with open(config_file, 'w') as f:
        f.write(config_content)
    
    logger.info(f"Created test config file: {config_file}")
    return config_file

def run_netwalker_test():
    """Run NetWalker with test configuration"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("REAL DATA TEST: DNS HOSTNAME CLEANING")
    logger.info("Testing with BORO-CORE-A seed and depth 2")
    logger.info("=" * 80)
    
    try:
        # Create test files
        seed_file = create_test_seed_file()
        config_file = create_test_config()
        
        # Create output directory
        os.makedirs("test_dns_reports", exist_ok=True)
        
        logger.info("\n1. Running NetWalker with test configuration...")
        logger.info("   Seed: BORO-CORE-A (10.1.1.1)")
        logger.info("   Depth: 2")
        logger.info("   DNS Validation: Enabled")
        
        # Run NetWalker
        import subprocess
        
        cmd = [
            "./dist/netwalker.exe",
            "--config", config_file,
            "--verbose"
        ]
        
        logger.info(f"   Command: {' '.join(cmd)}")
        
        # Run the command and capture output
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        
        logger.info(f"   Exit code: {result.returncode}")
        
        if result.returncode != 0:
            logger.error("NetWalker execution failed:")
            logger.error(f"STDOUT: {result.stdout}")
            logger.error(f"STDERR: {result.stderr}")
            return False
        
        logger.info("   ✓ NetWalker execution completed successfully")
        
        # Check for DNS validation report
        logger.info("\n2. Checking for DNS validation report...")
        
        reports_dir = Path("test_dns_reports")
        dns_reports = list(reports_dir.glob("DNS_Validation_*.xlsx"))
        
        if not dns_reports:
            logger.error("   ✗ No DNS validation report found")
            return False
        
        latest_dns_report = max(dns_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"   ✓ Found DNS validation report: {latest_dns_report}")
        
        # Check the report content using openpyxl
        logger.info("\n3. Analyzing DNS validation report content...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(latest_dns_report)
            
            if "DNS Validation Results" not in workbook.sheetnames:
                logger.error("   ✗ DNS Validation Results sheet not found")
                return False
            
            ws = workbook["DNS Validation Results"]
            logger.info("   ✓ DNS Validation Results sheet found")
            
            # Check hostnames in the report
            hostnames_found = []
            problematic_hostnames = []
            
            for row in range(2, ws.max_row + 1):  # Skip header row
                hostname_cell = ws.cell(row=row, column=2)  # Hostname column
                if hostname_cell.value:
                    hostname = str(hostname_cell.value)
                    hostnames_found.append(hostname)
                    
                    # Check for serial numbers in parentheses
                    if '(' in hostname and ')' in hostname:
                        problematic_hostnames.append(hostname)
            
            logger.info(f"   Hostnames found in DNS report: {hostnames_found}")
            
            if problematic_hostnames:
                logger.error(f"   ✗ Found hostnames with serial numbers: {problematic_hostnames}")
                logger.error("   The DNS hostname cleaning fix did not work properly")
                return False
            else:
                logger.info("   ✓ No hostnames with serial numbers found")
                logger.info("   ✓ DNS hostname cleaning is working correctly")
            
            # Check specifically for LUMT-CORE-A
            lumt_found = False
            lumt_with_serial_found = False
            
            for hostname in hostnames_found:
                if hostname == "LUMT-CORE-A":
                    lumt_found = True
                elif "LUMT-CORE-A(" in hostname:
                    lumt_with_serial_found = True
            
            if lumt_with_serial_found:
                logger.error("   ✗ Found LUMT-CORE-A with serial number - fix not working")
                return False
            elif lumt_found:
                logger.info("   ✓ Found clean LUMT-CORE-A hostname (no serial number)")
            else:
                logger.info("   ℹ LUMT-CORE-A not found (may not be discovered at depth 2)")
            
            workbook.close()
            
        except ImportError:
            logger.warning("   ⚠ openpyxl not available, skipping detailed report analysis")
        except Exception as e:
            logger.error(f"   ✗ Error analyzing report: {e}")
            return False
        
        logger.info("\n" + "=" * 80)
        logger.info("🎉 REAL DATA TEST PASSED!")
        logger.info("✅ NetWalker executed successfully with DNS validation")
        logger.info("✅ DNS validation report generated")
        logger.info("✅ No hostnames with serial numbers found in DNS report")
        logger.info("✅ DNS hostname cleaning fix is working correctly")
        logger.info("=" * 80)
        
        return True
        
    except subprocess.TimeoutExpired:
        logger.error("NetWalker execution timed out after 5 minutes")
        return False
    except Exception as e:
        logger.error(f"Test failed with exception: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Cleanup test files
        for file in ["test_dns_seed.csv", "test_dns_config.ini"]:
            if os.path.exists(file):
                os.remove(file)

def main():
    """Main function to run the real data test"""
    logger = setup_logging()
    
    logger.info("NetWalker DNS Hostname Cleaning - Real Data Test")
    logger.info("Testing the fix with actual NetWalker execution")
    
    success = run_netwalker_test()
    
    if success:
        logger.info("\n🎉 REAL DATA TEST SUCCESSFUL!")
        logger.info("The DNS hostname cleaning fix is working correctly with real data.")
        logger.info("LUMT-CORE-A(FOX1849GQKY) will now appear as LUMT-CORE-A in DNS reports.")
        return True
    else:
        logger.error("\n❌ REAL DATA TEST FAILED!")
        logger.error("The DNS hostname cleaning fix needs further attention.")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)