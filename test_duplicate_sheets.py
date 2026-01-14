#!/usr/bin/env python3
"""
Test script to investigate duplicate sheets in seed workbooks.

This script analyzes the WLTX-CORE-A seed workbook to understand
why there might be duplicate sheets for some devices.
"""

import sys
import os
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def analyze_seed_workbook():
    """Analyze the WLTX-CORE-A seed workbook for duplicate sheets"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("ANALYZING SEED WORKBOOK FOR DUPLICATE SHEETS")
    logger.info("Looking for WLTX-CORE-A seed workbook")
    logger.info("=" * 80)
    
    try:
        # Find WLTX-CORE-A seed workbook
        reports_dir = Path("reports")
        wltx_reports = list(reports_dir.glob("Seed_WLTX-CORE-A_*.xlsx"))
        
        if not wltx_reports:
            logger.error("No WLTX-CORE-A seed reports found")
            logger.info("Available seed reports:")
            seed_reports = list(reports_dir.glob("Seed_*.xlsx"))
            for report in seed_reports[-10:]:  # Show last 10
                logger.info(f"  {report.name}")
            return False
        
        latest_wltx_report = max(wltx_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"Latest WLTX seed report: {latest_wltx_report}")
        
        # Load and analyze the workbook
        from openpyxl import load_workbook
        
        workbook = load_workbook(latest_wltx_report)
        sheet_names = workbook.sheetnames
        
        logger.info(f"\nTotal sheets in workbook: {len(sheet_names)}")
        logger.info("Sheet names:")
        
        # Categorize sheets
        neighbor_sheets = []
        other_sheets = []
        
        for sheet_name in sheet_names:
            logger.info(f"  {sheet_name}")
            if sheet_name.startswith("Neighbors_"):
                neighbor_sheets.append(sheet_name)
            else:
                other_sheets.append(sheet_name)
        
        logger.info(f"\nSheet categories:")
        logger.info(f"  Other sheets: {len(other_sheets)} - {other_sheets}")
        logger.info(f"  Neighbor sheets: {len(neighbor_sheets)}")
        
        # Look for potential duplicates in neighbor sheets
        logger.info(f"\nAnalyzing neighbor sheets for duplicates:")
        
        # Extract device names from neighbor sheet names
        device_names = []
        for sheet_name in neighbor_sheets:
            if sheet_name.startswith("Neighbors_"):
                device_name = sheet_name.replace("Neighbors_", "")
                device_names.append(device_name)
        
        # Check for duplicates
        unique_devices = set(device_names)
        duplicate_devices = []
        
        for device in unique_devices:
            count = device_names.count(device)
            if count > 1:
                duplicate_devices.append((device, count))
        
        if duplicate_devices:
            logger.error(f"\n❌ DUPLICATE SHEETS FOUND!")
            logger.error(f"Found {len(duplicate_devices)} devices with duplicate sheets:")
            for device, count in duplicate_devices:
                logger.error(f"  {device}: {count} sheets")
                
                # Show the actual sheet names for this device
                matching_sheets = [s for s in neighbor_sheets if device in s]
                for sheet in matching_sheets:
                    logger.error(f"    - {sheet}")
        else:
            logger.info(f"\n✅ NO DUPLICATE SHEETS FOUND")
            logger.info(f"All {len(neighbor_sheets)} neighbor sheets are unique")
        
        # Analyze sheet content for potential issues
        logger.info(f"\nAnalyzing sheet content for hostname variations...")
        
        hostname_variations = {}
        
        for sheet_name in neighbor_sheets:
            try:
                ws = workbook[sheet_name]
                
                # Get the device name from the header (A1 cell)
                header_cell = ws['A1']
                if header_cell.value:
                    header_text = str(header_cell.value)
                    if "Neighbor Details for " in header_text:
                        device_name = header_text.replace("Neighbor Details for ", "")
                        
                        # Track variations
                        sheet_device = sheet_name.replace("Neighbors_", "")
                        if device_name != sheet_device:
                            if device_name not in hostname_variations:
                                hostname_variations[device_name] = []
                            hostname_variations[device_name].append(sheet_name)
                            
            except Exception as e:
                logger.warning(f"Could not analyze sheet {sheet_name}: {e}")
        
        if hostname_variations:
            logger.warning(f"\n⚠️ HOSTNAME VARIATIONS FOUND:")
            for device_name, sheets in hostname_variations.items():
                logger.warning(f"  Device '{device_name}' appears in sheets:")
                for sheet in sheets:
                    logger.warning(f"    - {sheet}")
        else:
            logger.info(f"\n✅ NO HOSTNAME VARIATIONS FOUND")
        
        workbook.close()
        
        # Summary
        logger.info(f"\n" + "=" * 80)
        logger.info(f"ANALYSIS SUMMARY")
        logger.info(f"=" * 80)
        logger.info(f"Total sheets: {len(sheet_names)}")
        logger.info(f"Neighbor sheets: {len(neighbor_sheets)}")
        logger.info(f"Unique devices: {len(unique_devices)}")
        logger.info(f"Duplicate devices: {len(duplicate_devices)}")
        logger.info(f"Hostname variations: {len(hostname_variations)}")
        
        if duplicate_devices or hostname_variations:
            logger.error(f"\n❌ ISSUES FOUND IN SEED WORKBOOK")
            logger.error(f"There are duplicate sheets or hostname variations that need investigation")
            return False
        else:
            logger.info(f"\n✅ SEED WORKBOOK LOOKS GOOD")
            logger.info(f"No duplicate sheets or hostname variations found")
            return True
        
    except ImportError:
        logger.error("openpyxl not available - cannot analyze Excel workbook")
        return False
    except Exception as e:
        logger.error(f"Error analyzing seed workbook: {e}")
        import traceback
        traceback.print_exc()
        return False

def run_wltx_test():
    """Run a test with WLTX-CORE-A to reproduce the issue"""
    logger = setup_logging()
    
    logger.info("\n" + "=" * 80)
    logger.info("RUNNING WLTX-CORE-A TEST TO REPRODUCE DUPLICATE SHEETS")
    logger.info("=" * 80)
    
    try:
        # Create test seed file with WLTX-CORE-A
        import csv
        
        seed_file = "test_wltx_seed.csv"
        
        with open(seed_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['hostname', 'ip_address', 'status'])
            writer.writerow(['WLTX-CORE-A', '10.1.1.1', ''])  # Use placeholder IP
        
        logger.info(f"Created test seed file: {seed_file}")
        
        # Create test config
        config_file = "test_wltx_config.ini"
        
        config_content = """[discovery]
seed_file = test_wltx_seed.csv
max_depth = 2
timeout = 30
concurrent_limit = 5

[output]
site_boundary_pattern = *-CORE-*
output_directory = test_wltx_reports
create_master_inventory = true
create_site_reports = true

[credentials]
username = admin
password = admin
enable_password = admin

[filtering]
enable_device_filtering = false
enable_lumt_filtering = false

[validation]
enable_dns_validation = false
"""
        
        with open(config_file, 'w') as f:
            f.write(config_content)
        
        logger.info(f"Created test config file: {config_file}")
        
        # Create output directory
        os.makedirs("test_wltx_reports", exist_ok=True)
        
        logger.info("Running NetWalker with WLTX-CORE-A...")
        
        # Run NetWalker
        import subprocess
        
        cmd = [
            "./dist/netwalker.exe",
            "--config", config_file,
            "--verbose"
        ]
        
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        
        logger.info(f"NetWalker exit code: {result.returncode}")
        
        if result.returncode != 0:
            logger.error("NetWalker execution failed:")
            logger.error(f"STDOUT: {result.stdout}")
            logger.error(f"STDERR: {result.stderr}")
            return False
        
        logger.info("✅ NetWalker execution completed successfully")
        
        # Analyze the generated seed workbook
        reports_dir = Path("test_wltx_reports")
        wltx_reports = list(reports_dir.glob("Seed_WLTX-CORE-A_*.xlsx"))
        
        if not wltx_reports:
            logger.error("No WLTX seed report generated")
            return False
        
        latest_report = max(wltx_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"Generated WLTX seed report: {latest_report}")
        
        # Analyze for duplicates
        from openpyxl import load_workbook
        
        workbook = load_workbook(latest_report)
        sheet_names = workbook.sheetnames
        
        logger.info(f"Sheets in generated workbook: {len(sheet_names)}")
        
        neighbor_sheets = [s for s in sheet_names if s.startswith("Neighbors_")]
        logger.info(f"Neighbor sheets: {len(neighbor_sheets)}")
        
        for sheet in neighbor_sheets:
            logger.info(f"  {sheet}")
        
        # Check for duplicates
        device_names = [s.replace("Neighbors_", "") for s in neighbor_sheets]
        unique_devices = set(device_names)
        
        duplicates = []
        for device in unique_devices:
            count = device_names.count(device)
            if count > 1:
                duplicates.append((device, count))
        
        if duplicates:
            logger.error(f"\n❌ DUPLICATE SHEETS REPRODUCED!")
            for device, count in duplicates:
                logger.error(f"  {device}: {count} sheets")
        else:
            logger.info(f"\n✅ NO DUPLICATES IN TEST RUN")
        
        workbook.close()
        
        return len(duplicates) == 0
        
    except Exception as e:
        logger.error(f"Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Cleanup
        for file in ["test_wltx_seed.csv", "test_wltx_config.ini"]:
            if os.path.exists(file):
                os.remove(file)

def main():
    """Main function"""
    logger = setup_logging()
    
    logger.info("NetWalker Duplicate Sheets Investigation")
    logger.info("Analyzing WLTX-CORE-A seed workbook for duplicate sheets")
    
    # First, analyze existing reports
    existing_analysis = analyze_seed_workbook()
    
    # If we found issues or no existing reports, run a test
    if not existing_analysis:
        logger.info("\nRunning test to reproduce the issue...")
        test_result = run_wltx_test()
        
        if not test_result:
            logger.error("\n❌ DUPLICATE SHEETS ISSUE CONFIRMED")
            logger.error("The seed workbook generation has a bug causing duplicate sheets")
            return False
    
    logger.info("\n🎯 INVESTIGATION COMPLETE")
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)