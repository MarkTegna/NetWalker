"""
Test stack member expansion in Device Inventory sheet
"""

import sys
sys.path.insert(0, '.')

from openpyxl import Workbook
from netwalker.reports.excel_generator import ExcelReportGenerator
from netwalker.connection.data_models import StackMemberInfo
from datetime import datetime

# Create mock inventory with a VSS stack
inventory = {
    "10.1.1.1": {
        "hostname": "KGW-CORE-A",
        "primary_ip": "10.1.1.1",
        "platform": "IOS-XE",
        "software_version": "03.11.03.E",
        "vtp_version": "2",
        "serial_number": "JAE240213DA",  # This will be from Switch 1
        "hardware_model": "WS-C4500X-32",
        "uptime": "5 days, 3 hours",
        "discovery_depth": 0,
        "discovery_method": "Seed",
        "parent_device": "",
        "discovery_timestamp": datetime.now().isoformat(),
        "connection_method": "SSH",
        "status": "success",
        "connection_status": "success",
        "is_stack": True,
        "stack_members": [
            StackMemberInfo(
                switch_number=1,
                role="Active",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE240213DA",
                mac_address=None,
                software_version="03.11.03.E",
                state="Ok"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Standby",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE171504NJ",
                mac_address=None,
                software_version="03.11.03.E",
                state="Ok"
            )
        ]
    },
    "10.1.1.2": {
        "hostname": "SWITCH-01",
        "primary_ip": "10.1.1.2",
        "platform": "IOS",
        "software_version": "15.2(4)E",
        "vtp_version": "2",
        "serial_number": "FOC123456AB",
        "hardware_model": "WS-C2960X-48FPS-L",
        "uptime": "10 days, 2 hours",
        "discovery_depth": 1,
        "discovery_method": "CDP",
        "parent_device": "KGW-CORE-A",
        "discovery_timestamp": datetime.now().isoformat(),
        "connection_method": "SSH",
        "status": "success",
        "connection_status": "success",
        "is_stack": False,
        "stack_members": []
    }
}

print("=" * 80)
print("Stack Member Expansion Test")
print("=" * 80)

# Create Excel generator
config = {
    'output': {
        'reports_directory': './reports',
        'excel_format': 'xlsx'
    }
}

generator = ExcelReportGenerator(config)

# Create workbook and add device inventory sheet
workbook = Workbook()
workbook.remove(workbook.active)  # Remove default sheet

print("\nGenerating Device Inventory sheet with stack expansion...")
generator._create_device_inventory_sheet(workbook, inventory)

# Verify the sheet
ws = workbook["Device Inventory"]

print(f"\nSheet created with {ws.max_row} rows (including header)")

# Check headers
print("\nHeaders:")
headers = [cell.value for cell in ws[1]]
print(f"  Columns: {len(headers)}")
print(f"  Stack columns: Is Stack (col 9), Stack Role (col 10), Stack Number (col 11), Physical Device (col 12)")

# Check data rows
print("\nData rows:")
for row_idx in range(2, ws.max_row + 1):
    device_key = ws.cell(row=row_idx, column=1).value
    hostname = ws.cell(row=row_idx, column=2).value
    serial = ws.cell(row=row_idx, column=7).value
    is_stack = ws.cell(row=row_idx, column=9).value
    stack_role = ws.cell(row=row_idx, column=10).value
    stack_num = ws.cell(row=row_idx, column=11).value
    physical = ws.cell(row=row_idx, column=12).value
    
    print(f"  Row {row_idx}: {hostname:20s} | Serial: {serial:15s} | Stack: {is_stack:3s} | Role: {str(stack_role or 'N/A'):10s} | Num: {str(stack_num or 'N/A'):3s} | Physical: {physical}")

# Validate results
print("\n" + "=" * 80)
print("Validation:")
print("=" * 80)

expected_rows = 5  # Header + KGW-CORE-A (logical) + SW1 + SW2 + SWITCH-01
actual_rows = ws.max_row

if actual_rows == expected_rows:
    print(f"✓ Row count correct: {actual_rows} rows")
else:
    print(f"✗ Row count incorrect: expected {expected_rows}, got {actual_rows}")
    sys.exit(1)

# Check KGW-CORE-A logical device (row 2)
if ws.cell(row=2, column=2).value == "KGW-CORE-A":
    if ws.cell(row=2, column=9).value == "Yes" and ws.cell(row=2, column=12).value == "No":
        print("✓ KGW-CORE-A logical device row correct")
    else:
        print("✗ KGW-CORE-A logical device flags incorrect")
        sys.exit(1)
else:
    print("✗ KGW-CORE-A not found in row 2")
    sys.exit(1)

# Check SW1 (row 3)
if ws.cell(row=3, column=2).value == "KGW-CORE-A-SW1":
    serial = ws.cell(row=3, column=7).value
    role = ws.cell(row=3, column=10).value
    num = ws.cell(row=3, column=11).value
    physical = ws.cell(row=3, column=12).value
    
    if serial == "JAE240213DA" and role == "Active" and num == 1 and physical == "Yes":
        print("✓ KGW-CORE-A-SW1 physical device row correct")
    else:
        print(f"✗ KGW-CORE-A-SW1 data incorrect: serial={serial}, role={role}, num={num}, physical={physical}")
        sys.exit(1)
else:
    print("✗ KGW-CORE-A-SW1 not found in row 3")
    sys.exit(1)

# Check SW2 (row 4)
if ws.cell(row=4, column=2).value == "KGW-CORE-A-SW2":
    serial = ws.cell(row=4, column=7).value
    role = ws.cell(row=4, column=10).value
    num = ws.cell(row=4, column=11).value
    physical = ws.cell(row=4, column=12).value
    
    if serial == "JAE171504NJ" and role == "Standby" and num == 2 and physical == "Yes":
        print("✓ KGW-CORE-A-SW2 physical device row correct")
    else:
        print(f"✗ KGW-CORE-A-SW2 data incorrect: serial={serial}, role={role}, num={num}, physical={physical}")
        sys.exit(1)
else:
    print("✗ KGW-CORE-A-SW2 not found in row 4")
    sys.exit(1)

# Check SWITCH-01 (row 5)
if ws.cell(row=5, column=2).value == "SWITCH-01":
    is_stack = ws.cell(row=5, column=9).value
    physical = ws.cell(row=5, column=12).value
    
    if is_stack == "No" and physical == "No":
        print("✓ SWITCH-01 standalone device row correct")
    else:
        print(f"✗ SWITCH-01 flags incorrect: is_stack={is_stack}, physical={physical}")
        sys.exit(1)
else:
    print("✗ SWITCH-01 not found in row 5")
    sys.exit(1)

print("\n" + "=" * 80)
print("ALL TESTS PASSED!")
print("=" * 80)
print("\nStack member expansion is working correctly:")
print("  ✓ Logical stack device row created")
print("  ✓ Physical member rows created for each switch")
print("  ✓ Serial numbers correctly assigned to physical devices")
print("  ✓ Stack role and number populated")
print("  ✓ Physical Device flag correctly set")
print("  ✓ Standalone devices unaffected")

# Save for manual inspection
output_file = "test_stack_inventory.xlsx"
workbook.save(output_file)
print(f"\nTest workbook saved to: {output_file}")
