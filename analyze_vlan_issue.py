import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('reports/Discovery_BORO_20260112-22-03.xlsx')

# Check Device Inventory
print("=" * 80)
print("DEVICE INVENTORY - BORO-MDF devices")
print("=" * 80)
ws_inv = wb['Device Inventory']
headers_inv = [cell.value for cell in ws_inv[1]]
print(f"Headers: {headers_inv}\n")

for row in range(2, ws_inv.max_row + 1):
    hostname = ws_inv.cell(row, 1).value
    if hostname and 'BORO-MDF' in str(hostname):
        row_data = [ws_inv.cell(row, col).value for col in range(1, len(headers_inv) + 1)]
        print(f"{hostname}:")
        for i, header in enumerate(headers_inv):
            print(f"  {header}: {row_data[i]}")
        print()

# Check VLAN Inventory
print("\n" + "=" * 80)
print("VLAN INVENTORY - BORO-MDF devices")
print("=" * 80)
ws_vlan = wb['VLAN Inventory']
headers_vlan = [cell.value for cell in ws_vlan[1]]
print(f"Headers: {headers_vlan}\n")

# Group VLANs by device
device_vlans = {}
for row in range(2, ws_vlan.max_row + 1):
    device_name = ws_vlan.cell(row, 1).value
    if device_name and 'BORO-MDF' in str(device_name):
        if device_name not in device_vlans:
            device_vlans[device_name] = []
        vlan_id = ws_vlan.cell(row, 2).value
        vlan_name = ws_vlan.cell(row, 3).value
        device_vlans[device_name].append((vlan_id, vlan_name))

for device in sorted(device_vlans.keys()):
    print(f"{device}:")
    for vlan_id, vlan_name in sorted(device_vlans[device]):
        print(f"  VLAN {vlan_id}: {vlan_name}")
    print()

# Check which devices are missing
all_mdf_devices = ['BORO-MDF-SW01', 'BORO-MDF-SW02', 'BORO-MDF-SW03', 'BORO-MDF-SW04', 'BORO-MDF-SW05']
print("\n" + "=" * 80)
print("MISSING VLAN DATA")
print("=" * 80)
for device in all_mdf_devices:
    if device not in device_vlans:
        print(f"{device}: NO VLAN DATA COLLECTED")
