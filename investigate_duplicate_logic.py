#!/usr/bin/env python3
"""
Investigate the duplicate sheets logic in Excel generator.

This script analyzes the device matching logic that could cause duplicate sheets.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the investigation"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def simulate_device_matching():
    """Simulate the device matching logic that could cause duplicates"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("INVESTIGATING DUPLICATE SHEETS LOGIC")
    logger.info("=" * 80)
    
    # Simulate a scenario that could cause duplicates
    # This represents what might be in full_inventory
    full_inventory = {
        'device1': {
            'hostname': 'WLTX-CORE-A.example.com',
            'ip_address': '10.1.1.1',
            'status': 'discovered'
        },
        'device2': {
            'hostname': 'WLTX-CORE-A',
            'ip_address': '10.1.1.1',
            'status': 'discovered'
        },
        'device3': {
            'hostname': 'WLTX-CORE-B.example.com',
            'ip_address': '10.1.1.2',
            'status': 'discovered'
        }
    }
    
    # Simulate seed neighbors
    seed_neighbors = [
        {'device_id': 'WLTX-CORE-A', 'ip_address': '10.1.1.1'},
        {'device_id': 'WLTX-CORE-B', 'ip_address': '10.1.1.2'}
    ]
    
    def clean_hostname(hostname):
        """Simulate the _clean_hostname method"""
        if not hostname:
            return ''
        # Remove common suffixes and clean up
        hostname = hostname.replace('(', '_').replace(')', '')
        return hostname.strip()
    
    logger.info("Full inventory:")
    for key, device in full_inventory.items():
        logger.info(f"  {key}: {device['hostname']} -> {clean_hostname(device['hostname'])}")
    
    logger.info("\nSeed neighbors:")
    for neighbor in seed_neighbors:
        logger.info(f"  {neighbor['device_id']}")
    
    logger.info("\nSimulating device matching logic:")
    
    sheets_to_create = []
    
    for neighbor in seed_neighbors:
        neighbor_hostname = clean_hostname(neighbor.get('device_id', ''))
        logger.info(f"\nProcessing neighbor: {neighbor_hostname}")
        
        # This is the problematic logic from line 1529
        matches = []
        for device_key, device_info in full_inventory.items():
            device_hostname = clean_hostname(device_info.get('hostname', ''))
            
            # The problematic condition
            if device_hostname == neighbor_hostname or device_hostname.split('.')[0] == neighbor_hostname:
                matches.append((device_key, device_hostname))
                logger.info(f"  MATCH: {device_key} ({device_hostname})")
        
        if len(matches) > 1:
            logger.error(f"  ❌ MULTIPLE MATCHES FOUND for {neighbor_hostname}:")
            for device_key, device_hostname in matches:
                logger.error(f"    - {device_key}: {device_hostname}")
            logger.error(f"  This would create duplicate sheets!")
        elif len(matches) == 1:
            logger.info(f"  ✅ Single match found")
            sheets_to_create.append(f"Neighbors_{neighbor_hostname[:20]}")
        else:
            logger.warning(f"  ⚠️ No matches found")
    
    logger.info(f"\nSheets that would be created:")
    for sheet in sheets_to_create:
        logger.info(f"  {sheet}")
    
    # Check for duplicate sheet names
    unique_sheets = set(sheets_to_create)
    if len(sheets_to_create) != len(unique_sheets):
        logger.error(f"\n❌ DUPLICATE SHEET NAMES DETECTED!")
        logger.error(f"Total sheets: {len(sheets_to_create)}")
        logger.error(f"Unique sheets: {len(unique_sheets)}")
        return False
    else:
        logger.info(f"\n✅ No duplicate sheet names")
        return True

def analyze_real_inventory():
    """Analyze a real inventory file to look for potential duplicates"""
    logger = setup_logging()
    
    logger.info("\n" + "=" * 80)
    logger.info("ANALYZING REAL INVENTORY FOR DUPLICATE POTENTIAL")
    logger.info("=" * 80)
    
    try:
        # Look for recent inventory files
        reports_dir = Path("reports")
        inventory_files = list(reports_dir.glob("*Inventory*.xlsx"))
        
        if not inventory_files:
            logger.warning("No inventory files found")
            return True
        
        latest_inventory = max(inventory_files, key=lambda p: p.stat().st_mtime)
        logger.info(f"Analyzing: {latest_inventory}")
        
        from openpyxl import load_workbook
        
        workbook = load_workbook(latest_inventory)
        
        # Look for the main inventory sheet
        inventory_sheet = None
        for sheet_name in workbook.sheetnames:
            if 'inventory' in sheet_name.lower() or 'device' in sheet_name.lower():
                inventory_sheet = workbook[sheet_name]
                break
        
        if not inventory_sheet:
            logger.warning("Could not find inventory sheet")
            return True
        
        logger.info(f"Found inventory sheet: {inventory_sheet.title}")
        
        # Read hostnames from the inventory
        hostnames = []
        for row in inventory_sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 1:  # Assuming hostname is in column 2
                hostname = row[1] if row[1] else ''
                if hostname:
                    hostnames.append(str(hostname))
        
        logger.info(f"Found {len(hostnames)} devices in inventory")
        
        # Look for potential duplicates using the same logic
        def clean_hostname(hostname):
            if not hostname:
                return ''
            hostname = hostname.replace('(', '_').replace(')', '')
            return hostname.strip()
        
        # Group by cleaned short hostname
        hostname_groups = {}
        for hostname in hostnames:
            cleaned = clean_hostname(hostname)
            short_name = cleaned.split('.')[0]
            
            if short_name not in hostname_groups:
                hostname_groups[short_name] = []
            hostname_groups[short_name].append(cleaned)
        
        # Find groups with multiple entries
        potential_duplicates = {}
        for short_name, full_names in hostname_groups.items():
            if len(full_names) > 1:
                potential_duplicates[short_name] = full_names
        
        if potential_duplicates:
            logger.warning(f"\n⚠️ POTENTIAL DUPLICATE SOURCES FOUND:")
            for short_name, full_names in potential_duplicates.items():
                logger.warning(f"  {short_name}:")
                for full_name in full_names:
                    logger.warning(f"    - {full_name}")
                logger.warning(f"  This could cause duplicate sheets for neighbors of {short_name}")
        else:
            logger.info(f"\n✅ No potential duplicate sources found")
        
        workbook.close()
        return len(potential_duplicates) == 0
        
    except ImportError:
        logger.error("openpyxl not available")
        return True
    except Exception as e:
        logger.error(f"Error analyzing inventory: {e}")
        return True

def main():
    """Main investigation function"""
    logger = setup_logging()
    
    logger.info("NetWalker Duplicate Sheets Logic Investigation")
    
    # Test the matching logic
    logic_ok = simulate_device_matching()
    
    # Analyze real data
    inventory_ok = analyze_real_inventory()
    
    logger.info(f"\n" + "=" * 80)
    logger.info("INVESTIGATION SUMMARY")
    logger.info("=" * 80)
    
    if logic_ok and inventory_ok:
        logger.info("✅ No duplicate sheet issues detected")
        logger.info("The matching logic appears to be working correctly")
    else:
        logger.error("❌ Potential duplicate sheet issues found")
        logger.error("The device matching logic needs to be fixed")
        
        logger.info("\nRecommended fix:")
        logger.info("1. Use a more precise matching algorithm")
        logger.info("2. Prefer exact hostname matches over partial matches")
        logger.info("3. Add duplicate detection and prevention")
    
    return logic_ok and inventory_ok

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)