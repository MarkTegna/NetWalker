#!/usr/bin/env python3
"""
Check the DNS validation report to verify hostname cleaning is working.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def check_latest_dns_report():
    """Check the latest DNS validation report for hostname cleaning"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("CHECKING LATEST DNS VALIDATION REPORT")
    logger.info("Verifying hostname cleaning is working correctly")
    logger.info("=" * 80)
    
    try:
        # Find the latest DNS validation report
        reports_dir = Path("reports")
        dns_reports = list(reports_dir.glob("DNS_Validation_*.xlsx"))
        
        if not dns_reports:
            logger.error("No DNS validation reports found")
            return False
        
        latest_dns_report = max(dns_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"Latest DNS report: {latest_dns_report}")
        
        # Load and analyze the report
        from openpyxl import load_workbook
        
        workbook = load_workbook(latest_dns_report)
        
        if "DNS Validation Results" not in workbook.sheetnames:
            logger.error("DNS Validation Results sheet not found")
            return False
        
        ws = workbook["DNS Validation Results"]
        logger.info("DNS Validation Results sheet found")
        
        # Analyze hostnames
        logger.info("\nHostnames found in DNS validation report:")
        
        hostnames_found = []
        problematic_hostnames = []
        clean_hostnames = []
        
        for row in range(2, ws.max_row + 1):  # Skip header row
            hostname_cell = ws.cell(row=row, column=2)  # Hostname column
            ip_cell = ws.cell(row=row, column=3)  # IP column
            
            if hostname_cell.value:
                hostname = str(hostname_cell.value)
                ip = str(ip_cell.value) if ip_cell.value else ""
                
                hostnames_found.append(hostname)
                logger.info(f"  {hostname} ({ip})")
                
                # Check for serial numbers in parentheses
                if '(' in hostname and ')' in hostname:
                    problematic_hostnames.append(hostname)
                    logger.error(f"    ✗ PROBLEM: Contains serial number in parentheses")
                else:
                    clean_hostnames.append(hostname)
                    logger.info(f"    ✓ CLEAN: No serial numbers")
        
        workbook.close()
        
        # Summary
        logger.info(f"\nSUMMARY:")
        logger.info(f"Total hostnames: {len(hostnames_found)}")
        logger.info(f"Clean hostnames: {len(clean_hostnames)}")
        logger.info(f"Problematic hostnames: {len(problematic_hostnames)}")
        
        if problematic_hostnames:
            logger.error(f"\n❌ HOSTNAME CLEANING FAILED!")
            logger.error(f"Found {len(problematic_hostnames)} hostnames with serial numbers:")
            for hostname in problematic_hostnames:
                logger.error(f"  - {hostname}")
            return False
        else:
            logger.info(f"\n🎉 HOSTNAME CLEANING SUCCESS!")
            logger.info(f"✅ All {len(hostnames_found)} hostnames are properly cleaned")
            logger.info(f"✅ No serial numbers in parentheses found")
            
            # Check specifically for LUMT-CORE-A
            lumt_variants = [h for h in hostnames_found if 'LUMT-CORE-A' in h]
            if lumt_variants:
                logger.info(f"✅ LUMT-CORE-A variants found: {lumt_variants}")
                if any('(' in variant for variant in lumt_variants):
                    logger.error("✗ LUMT-CORE-A still has serial numbers")
                    return False
                else:
                    logger.info("✅ LUMT-CORE-A properly cleaned (no serial numbers)")
            
            return True
        
    except ImportError:
        logger.error("openpyxl not available - cannot analyze Excel report")
        return False
    except Exception as e:
        logger.error(f"Error checking DNS report: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function"""
    logger = setup_logging()
    
    logger.info("NetWalker DNS Hostname Cleaning Verification")
    logger.info("Checking latest DNS validation report")
    
    success = check_latest_dns_report()
    
    if success:
        logger.info("\n🎉 VERIFICATION SUCCESSFUL!")
        logger.info("The DNS hostname cleaning fix is working correctly.")
        return True
    else:
        logger.error("\n❌ VERIFICATION FAILED!")
        logger.error("The DNS hostname cleaning fix needs attention.")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)