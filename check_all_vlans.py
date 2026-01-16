import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('reports/Discovery_BORO_20260112-22-03.xlsx')
ws_vlan = wb['VLAN Inventory']

# Get all VLANs for BORO-MDF-SW01
print("All VLANs on BORO-MDF-SW01:")
for row in range(2, ws_vlan.max_row + 1):
    device_name = ws_vlan.cell(row, 1).value
    if device_name == 'BORO-MDF-SW01':
        vlan_id = ws_vlan.cell(row, 2).value
        vlan_name = ws_vlan.cell(row, 3).value
        port_count = ws_vlan.cell(row, 4).value
        portchannel_count = ws_vlan.cell(row, 5).value
        print(f"  VLAN {vlan_id}: {vlan_name} (Ports: {port_count}, PortChannels: {portchannel_count})")

print("\n" + "=" * 80)
print("Searching for VLAN 461 across ALL devices:")
print("=" * 80)
found_461 = False
for row in range(2, ws_vlan.max_row + 1):
    vlan_id = ws_vlan.cell(row, 2).value
    if vlan_id == 461:
        found_461 = True
        device_name = ws_vlan.cell(row, 1).value
        vlan_name = ws_vlan.cell(row, 3).value
        port_count = ws_vlan.cell(row, 4).value
        portchannel_count = ws_vlan.cell(row, 5).value
        print(f"{device_name}: VLAN 461 - {vlan_name} (Ports: {port_count}, PortChannels: {portchannel_count})")

if not found_461:
    print("VLAN 461 was NOT found on ANY device in this discovery!")
