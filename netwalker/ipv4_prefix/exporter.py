"""
Exporter Component for IPv4 Prefix Inventory Module

This module handles the export of prefix data to CSV, Excel, and database formats.

Author: Mark Oldham
"""

import logging
import csv
import os
from typing import List, Optional
from datetime import datetime

from netwalker.ipv4_prefix.data_models import (
    NormalizedPrefix, DeduplicatedPrefix, CollectionException, SummarizationRelationship
)


class CSVExporter:
    """Exports prefix data to CSV files."""
    
    def __init__(self):
        """Initialize CSV exporter."""
        self.logger = logging.getLogger(__name__)
    
    def export_prefixes(self, prefixes: List[NormalizedPrefix], output_dir: str) -> str:
        """
        Export to prefixes.csv.
        
        Columns: device, platform, vrf, prefix, source, protocol, timestamp
        Sort order: vrf, prefix, device
        
        Args:
            prefixes: List of NormalizedPrefix objects
            output_dir: Output directory path
            
        Returns:
            Path to created CSV file
            
        Requirements:
            - 10.1: Create prefixes.csv file
            - 10.2: Include specified columns
            - 10.3: Sort by vrf, prefix, device
            - 10.4: Use UTF-8 encoding
            - 10.5: Log output file path
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Build output file path
        output_file = os.path.join(output_dir, 'prefixes.csv')
        
        # Sort prefixes by vrf, prefix, device
        sorted_prefixes = sorted(prefixes, key=lambda p: (p.vrf, p.prefix, p.device))
        
        # Write CSV file
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(['device', 'platform', 'vrf', 'prefix', 'source', 'protocol', 'vlan', 'interface', 'timestamp'])
            
            # Write data rows
            for prefix in sorted_prefixes:
                writer.writerow([
                    prefix.device,
                    prefix.platform,
                    prefix.vrf,
                    prefix.prefix,
                    prefix.source,
                    prefix.protocol,
                    prefix.vlan if prefix.vlan is not None else '',
                    prefix.interface if prefix.interface else '',
                    prefix.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self.logger.info(f"Exported {len(sorted_prefixes)} prefixes to: {output_file}")
        return output_file
    
    def export_deduplicated(self, prefixes: List[DeduplicatedPrefix], output_dir: str) -> str:
        """
        Export to prefixes_dedup_by_vrf.csv.
        
        Columns: vrf, prefix, device_count, device_list
        Sort order: vrf, prefix
        
        Args:
            prefixes: List of DeduplicatedPrefix objects
            output_dir: Output directory path
            
        Returns:
            Path to created CSV file
            
        Requirements:
            - 11.1: Create prefixes_dedup_by_vrf.csv file
            - 11.2: Include specified columns
            - 11.3: Sort by vrf, prefix
            - 11.4: Format device_list as semicolon-separated string
            - 11.5: Log output file path
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Build output file path
        output_file = os.path.join(output_dir, 'prefixes_dedup_by_vrf.csv')
        
        # Sort prefixes by vrf, prefix
        sorted_prefixes = sorted(prefixes, key=lambda p: (p.vrf, p.prefix))
        
        # Write CSV file
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(['vrf', 'prefix', 'device_count', 'device_list'])
            
            # Write data rows
            for prefix in sorted_prefixes:
                # Format device_list as semicolon-separated string
                device_list_str = ';'.join(prefix.device_list)
                
                writer.writerow([
                    prefix.vrf,
                    prefix.prefix,
                    prefix.device_count,
                    device_list_str
                ])
        
        self.logger.info(f"Exported {len(sorted_prefixes)} deduplicated prefixes to: {output_file}")
        return output_file
    
    def export_exceptions(self, exceptions: List[CollectionException], output_dir: str) -> str:
        """
        Export to exceptions.csv.
        
        Columns: device, command, error_type, raw_token, timestamp
        
        Args:
            exceptions: List of CollectionException objects
            output_dir: Output directory path
            
        Returns:
            Path to created CSV file
            
        Requirements:
            - 12.1: Create exceptions.csv file
            - 12.2: Include specified columns
            - 12.3: Record command failures
            - 12.4: Record unresolved prefixes
            - 12.5: Record parsing errors
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Build output file path
        output_file = os.path.join(output_dir, 'exceptions.csv')
        
        # Sort exceptions by device, timestamp
        sorted_exceptions = sorted(exceptions, key=lambda e: (e.device, e.timestamp))
        
        # Write CSV file
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(['device', 'command', 'error_type', 'raw_token', 'error_message', 'timestamp'])
            
            # Write data rows
            for exception in sorted_exceptions:
                writer.writerow([
                    exception.device,
                    exception.command,
                    exception.error_type,
                    exception.raw_token or '',
                    exception.error_message,
                    exception.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])
        
        self.logger.info(f"Exported {len(sorted_exceptions)} exceptions to: {output_file}")
        return output_file


class ExcelExporter:
    """Exports prefix data to Excel with formatting."""
    
    def __init__(self):
        """Initialize Excel exporter."""
        self.logger = logging.getLogger(__name__)
    
    def export(self, prefixes: List[NormalizedPrefix],
               deduplicated: List[DeduplicatedPrefix],
               exceptions: List[CollectionException],
               output_dir: str) -> str:
        """
        Export to Excel workbook with three sheets:
        - Prefixes: All collected prefixes
        - Deduplicated: Unique prefixes by VRF
        - Exceptions: Errors and unresolved prefixes
        
        Apply NetWalker formatting:
        - Header row: bold, colored background
        - Auto-sized columns
        - Data filters
        
        Args:
            prefixes: List of NormalizedPrefix objects
            deduplicated: List of DeduplicatedPrefix objects
            exceptions: List of CollectionException objects
            output_dir: Output directory path
            
        Returns:
            Path to created Excel file
            
        Requirements:
            - 13.1: Create workbook with three sheets
            - 13.2: Apply header formatting
            - 13.3: Apply column auto-sizing
            - 13.4: Apply data filters
            - 13.5: Use existing NetWalker Excel patterns
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.logger.error("openpyxl library not available - cannot create Excel export")
            return ""
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        # Build output file path with timestamp
        timestamp = datetime.now().strftime('%Y%m%d-%H%M')
        output_file = os.path.join(output_dir, f'ipv4_prefix_inventory_{timestamp}.xlsx')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Prefixes sheet
        self._create_prefixes_sheet(wb, prefixes)
        
        # Create Deduplicated sheet
        self._create_deduplicated_sheet(wb, deduplicated)
        
        # Create Exceptions sheet
        self._create_exceptions_sheet(wb, exceptions)
        
        # Save workbook
        wb.save(output_file)
        
        self.logger.info(f"Exported Excel workbook to: {output_file}")
        return output_file
    
    def _create_prefixes_sheet(self, wb, prefixes: List[NormalizedPrefix]):
        """Create Prefixes sheet with all collected prefixes."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        ws = wb.create_sheet('Prefixes')
        
        # Define headers
        headers = ['Device', 'Platform', 'VRF', 'Prefix', 'Source', 'Protocol', 'VLAN', 'Interface', 'Timestamp']
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Sort prefixes
        sorted_prefixes = sorted(prefixes, key=lambda p: (p.vrf, p.prefix, p.device))
        
        # Write data rows
        for row_num, prefix in enumerate(sorted_prefixes, 2):
            ws.cell(row=row_num, column=1, value=prefix.device)
            ws.cell(row=row_num, column=2, value=prefix.platform)
            ws.cell(row=row_num, column=3, value=prefix.vrf)
            ws.cell(row=row_num, column=4, value=prefix.prefix)
            ws.cell(row=row_num, column=5, value=prefix.source)
            ws.cell(row=row_num, column=6, value=prefix.protocol)
            ws.cell(row=row_num, column=7, value=prefix.vlan if prefix.vlan is not None else '')
            ws.cell(row=row_num, column=8, value=prefix.interface if prefix.interface else '')
            ws.cell(row=row_num, column=9, value=prefix.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].auto_size = True
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Apply filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_deduplicated_sheet(self, wb, deduplicated: List[DeduplicatedPrefix]):
        """Create Deduplicated sheet with unique prefixes by VRF."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        ws = wb.create_sheet('Deduplicated')
        
        # Define headers
        headers = ['VRF', 'Prefix', 'Device Count', 'Device List']
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Sort deduplicated prefixes
        sorted_prefixes = sorted(deduplicated, key=lambda p: (p.vrf, p.prefix))
        
        # Write data rows
        for row_num, prefix in enumerate(sorted_prefixes, 2):
            ws.cell(row=row_num, column=1, value=prefix.vrf)
            ws.cell(row=row_num, column=2, value=prefix.prefix)
            ws.cell(row=row_num, column=3, value=prefix.device_count)
            ws.cell(row=row_num, column=4, value=';'.join(prefix.device_list))
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].auto_size = True
            if col_num == 4:  # Device List column
                ws.column_dimensions[get_column_letter(col_num)].width = 40
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Apply filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_exceptions_sheet(self, wb, exceptions: List[CollectionException]):
        """Create Exceptions sheet with errors and unresolved prefixes."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        ws = wb.create_sheet('Exceptions')
        
        # Define headers
        headers = ['Device', 'Command', 'Error Type', 'Raw Token', 'Error Message', 'Timestamp']
        
        # Write headers with formatting
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Sort exceptions
        sorted_exceptions = sorted(exceptions, key=lambda e: (e.device, e.timestamp))
        
        # Write data rows
        for row_num, exception in enumerate(sorted_exceptions, 2):
            ws.cell(row=row_num, column=1, value=exception.device)
            ws.cell(row=row_num, column=2, value=exception.command)
            ws.cell(row=row_num, column=3, value=exception.error_type)
            ws.cell(row=row_num, column=4, value=exception.raw_token or '')
            ws.cell(row=row_num, column=5, value=exception.error_message)
            ws.cell(row=row_num, column=6, value=exception.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].auto_size = True
            if col_num == 5:  # Error Message column
                ws.column_dimensions[get_column_letter(col_num)].width = 50
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Apply filters
        ws.auto_filter.ref = ws.dimensions


class DatabaseExporter:
    """Exports prefix data to NetWalker database."""
    
    def __init__(self, db_manager):
        """
        Initialize database exporter.
        
        Args:
            db_manager: NetWalker DatabaseManager instance
        """
        self.db_manager = db_manager
        self.logger = logging.getLogger(__name__)
    
    def initialize_schema(self) -> bool:
        """
        Create tables if they don't exist:
        - ipv4_prefixes
        - ipv4_prefix_summarization
        
        Returns:
            True if schema initialization succeeded
            
        Requirements:
            - 14.1: Create ipv4_prefixes table
            - 15.1: Create ipv4_prefix_summarization table
        """
        try:
            # Check if database is enabled
            if not self.db_manager or not hasattr(self.db_manager, 'connection'):
                self.logger.info("Database not enabled - skipping schema initialization")
                return False
            
            # Call DatabaseManager's schema initialization method
            # This will be implemented in task 17.1
            if hasattr(self.db_manager, 'initialize_ipv4_prefix_schema'):
                return self.db_manager.initialize_ipv4_prefix_schema()
            else:
                self.logger.warning("DatabaseManager does not have initialize_ipv4_prefix_schema method")
                return False
                
        except Exception as e:
            self.logger.error(f"Failed to initialize database schema: {str(e)}")
            return False
    
    def upsert_prefix(self, prefix: NormalizedPrefix, device_id: int) -> Optional[int]:
        """
        Insert or update prefix record.
        
        Args:
            prefix: NormalizedPrefix object
            device_id: Device ID from devices table
            
        Returns:
            prefix_id if successful, None otherwise
            
        Requirements:
            - 14.2: Include specified columns
            - 14.3: Link to device via device_id foreign key
            - 14.4: Update last_seen for existing prefixes
            - 14.5: Insert with both timestamps for new prefixes
        """
        if not self.db_manager or not hasattr(self.db_manager, 'connection'):
            return None
        
        if not self.db_manager.is_connected():
            return None
        
        try:
            cursor = self.db_manager.connection.cursor()
            
            # Check if prefix exists
            cursor.execute("""
                SELECT prefix_id FROM ipv4_prefixes
                WHERE device_id = ? AND vrf = ? AND prefix = ? AND source = ?
            """, (device_id, prefix.vrf, prefix.prefix, prefix.source))
            
            row = cursor.fetchone()
            
            if row:
                # Update existing prefix - update last_seen
                prefix_id = row[0]
                cursor.execute("""
                    UPDATE ipv4_prefixes
                    SET last_seen = GETDATE(),
                        protocol = ?,
                        updated_at = GETDATE()
                    WHERE prefix_id = ?
                """, (prefix.protocol, prefix_id))
                
                self.logger.debug(f"Updated prefix {prefix.prefix} on device {device_id}")
            else:
                # Insert new prefix - set both first_seen and last_seen
                cursor.execute("""
                    INSERT INTO ipv4_prefixes
                    (device_id, vrf, prefix, source, protocol)
                    VALUES (?, ?, ?, ?, ?)
                """, (device_id, prefix.vrf, prefix.prefix, prefix.source, prefix.protocol))
                
                cursor.execute("SELECT @@IDENTITY")
                prefix_id = cursor.fetchone()[0]
                
                self.logger.debug(f"Inserted new prefix {prefix.prefix} on device {device_id}")
            
            self.db_manager.connection.commit()
            cursor.close()
            return prefix_id
            
        except Exception as e:
            self.logger.error(f"Error upserting prefix {prefix.prefix}: {str(e)}")
            if self.db_manager.connection:
                self.db_manager.connection.rollback()
            return None
    
    def upsert_summarization(self, relationship: SummarizationRelationship,
                            summary_prefix_id: int, component_prefix_id: int,
                            device_id: int) -> bool:
        """
        Insert or update summarization relationship.
        
        Links summary_prefix_id to component_prefix_id.
        
        Args:
            relationship: SummarizationRelationship object
            summary_prefix_id: ID of summary prefix
            component_prefix_id: ID of component prefix
            device_id: Device ID from devices table
            
        Returns:
            True if successful
            
        Requirements:
            - 15.2: Include specified columns
            - 15.4: Link summary to component via foreign keys
            - 15.5: Record device_id that performed summarization
        """
        if not self.db_manager or not hasattr(self.db_manager, 'connection'):
            return False
        
        if not self.db_manager.is_connected():
            return False
        
        try:
            cursor = self.db_manager.connection.cursor()
            
            # Check if relationship exists
            cursor.execute("""
                SELECT summarization_id FROM ipv4_prefix_summarization
                WHERE summary_prefix_id = ? AND component_prefix_id = ? AND device_id = ?
            """, (summary_prefix_id, component_prefix_id, device_id))
            
            row = cursor.fetchone()
            
            if not row:
                # Insert new relationship
                cursor.execute("""
                    INSERT INTO ipv4_prefix_summarization
                    (summary_prefix_id, component_prefix_id, device_id)
                    VALUES (?, ?, ?)
                """, (summary_prefix_id, component_prefix_id, device_id))
                
                self.logger.debug(f"Inserted summarization relationship: {relationship.summary_prefix} -> {relationship.component_prefix}")
            else:
                self.logger.debug(f"Summarization relationship already exists: {relationship.summary_prefix} -> {relationship.component_prefix}")
            
            self.db_manager.connection.commit()
            cursor.close()
            return True
            
        except Exception as e:
            self.logger.error(f"Error upserting summarization relationship: {str(e)}")
            if self.db_manager.connection:
                self.db_manager.connection.rollback()
            return False
