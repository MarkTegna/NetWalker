"""
Excel Exporter for Command Executor

This module exports command execution results to Excel files with professional
formatting including headers, auto-sized columns, and preserved line breaks.

Author: Mark Oldham
"""

import logging
import os
from datetime import datetime
from typing import List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .data_models import CommandResult

logger = logging.getLogger(__name__)


class CommandResultExporter:
    """
    Exports command execution results to Excel files.
    
    Features:
    - Timestamped filenames (Command_Results_YYYYMMDD-HH-MM.xlsx)
    - Professional header formatting (bold white on blue #366092)
    - Auto-adjusted column widths (max 100 characters)
    - Preserved line breaks in command output
    - Columns: Device Name, Device IP, Status, Command Output, Execution Time
    """
    
    def __init__(self, output_dir: str = '.'):
        """
        Initialize CommandResultExporter.
        
        Args:
            output_dir: Directory where Excel files will be saved (default: current directory)
        """
        self.output_dir = output_dir
        
        # Ensure output directory exists
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
        # Formatting styles matching NetWalker patterns
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        logger.info(f"CommandResultExporter initialized with output directory: {self.output_dir}")
    
    def export(self, results: List[CommandResult], command: str = "") -> str:
        """
        Export command results to an Excel file.
        
        Args:
            results: List of CommandResult objects to export
            command: The command that was executed (for reference)
            
        Returns:
            Path to the generated Excel file
            
        Raises:
            Exception: If Excel file creation or writing fails
        """
        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
        filename = f"Command_Results_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        logger.info(f"Exporting {len(results)} command results to: {filename}")
        
        try:
            # Create workbook
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Command Results"
            
            # Define headers
            headers = ["Device Name", "Device IP", "Status", "Command Output", "Execution Time"]
            
            # Write headers with formatting
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            
            # Write data rows
            for row_num, result in enumerate(results, start=2):
                ws.cell(row=row_num, column=1, value=result.device_name)
                ws.cell(row=row_num, column=2, value=result.ip_address)
                ws.cell(row=row_num, column=3, value=result.status)
                
                # Command output with wrap text to preserve line breaks
                output_cell = ws.cell(row=row_num, column=4, value=result.output)
                output_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Execution time formatted to 2 decimal places
                ws.cell(row=row_num, column=5, value=f"{result.execution_time:.2f}s")
            
            # Auto-adjust column widths (max 100 characters)
            self._auto_adjust_columns(ws, max_width=100)
            
            # Save workbook
            workbook.save(filepath)
            workbook.close()
            
            logger.info(f"Command results exported successfully to: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Failed to export command results to Excel: {e}")
            raise
    
    def _auto_adjust_columns(self, worksheet, max_width: int = 100):
        """
        Auto-adjust column widths based on content with a maximum width.
        
        Args:
            worksheet: The worksheet to adjust
            max_width: Maximum column width in characters (default: 100)
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Handle multi-line content by taking the longest line
                        cell_value = str(cell.value)
                        if '\n' in cell_value:
                            # For multi-line content, use the longest line
                            lines = cell_value.split('\n')
                            max_line_length = max(len(line) for line in lines)
                            if max_line_length > max_length:
                                max_length = max_line_length
                        else:
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                except:
                    pass
            
            # Add padding and cap at max_width
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logger.debug(f"Column {column_letter} width set to {adjusted_width}")
