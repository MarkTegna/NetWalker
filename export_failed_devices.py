"""Export all devices with status 'failed' to a spreadsheet."""
import pyodbc
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

server = "eit-prisqldb01.tgna.tegna.com"
database = "NetWalker"
username = "NetWalker"
password = base64.b64decode("Rmx1ZmZ5QnVubnlIaXRieWFCdXM=").decode()

conn_str = (
    f"DRIVER={{SQL Server}};"
    f"SERVER={server};"
    f"DATABASE={database};"
    f"UID={username};"
    f"PWD={password};"
    f"TrustServerCertificate=yes"
)

conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

# Check what status values exist
cursor.execute("SELECT DISTINCT status, COUNT(*) as cnt FROM devices GROUP BY status ORDER BY status")
print("Status counts:")
for row in cursor.fetchall():
    print(f"  {row.status}: {row.cnt}")

# Query failed devices
cursor.execute("""
    SELECT device_name, platform, hardware_model, capabilities, serial_number,
           connection_failures, first_seen, last_seen, status
    FROM devices
    WHERE status = 'failed'
    ORDER BY device_name
""")

rows = cursor.fetchall()
print(f"\nFound {len(rows)} failed devices")

if not rows:
    # Try other failure-like statuses
    cursor.execute("""
        SELECT device_name, platform, hardware_model, capabilities, serial_number,
               connection_failures, first_seen, last_seen, status
        FROM devices
        WHERE status NOT IN ('active', 'connected')
        ORDER BY device_name
    """)
    rows = cursor.fetchall()
    print(f"Found {len(rows)} non-active devices instead")

if rows:
    wb = Workbook()
    ws = wb.active
    ws.title = "Failed Devices"

    headers = ["Device Name", "Platform", "Hardware Model", "Capabilities", 
               "Serial Number", "Connection Failures", "First Seen", "Last Seen", "Status"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.device_name)
        ws.cell(row=row_idx, column=2, value=row.platform or "")
        ws.cell(row=row_idx, column=3, value=row.hardware_model or "")
        ws.cell(row=row_idx, column=4, value=row.capabilities or "")
        ws.cell(row=row_idx, column=5, value=row.serial_number or "")
        ws.cell(row=row_idx, column=6, value=row.connection_failures)
        ws.cell(row=row_idx, column=7, value=str(row.first_seen) if row.first_seen else "")
        ws.cell(row=row_idx, column=8, value=str(row.last_seen) if row.last_seen else "")
        ws.cell(row=row_idx, column=9, value=row.status or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 25

    timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
    filename = f"Failed_Devices_{timestamp}.xlsx"
    wb.save(filename)
    print(f"Exported to {filename}")
else:
    print("No failed devices found.")

cursor.close()
conn.close()
