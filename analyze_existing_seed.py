#!/usr/bin/env python3
"""
Analyze existing seed reports for duplicate sheets issue.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def analyze_seed_report(report_path):
    """Analyze a specific seed report for duplicate sheets"""
    logger = logging.getLogger(__name__)
    
    logger.info(f"Analyzing seed report: {report_path.name}")
    
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(report_path)
        sheet_names = workbook.sheetnames
        
        logger.info(f"Total sheets: {len(sheet_names)}")
        
        # Categorize sheets
        neighbor_sheets = [s for s in sheet_names if s.startswith("Neighbors_")]
        other_sheets = [s for s in sheet_names if not s.startswith("Neighbors_")]
        
        logger.info(f"Other sheets ({len(other_sheets)}): {other_sheets}")
        logger.info(f"Neighbor sheets ({len(neighbor_sheets)}):")
        
        # Extract device names and look for duplicates
        device_names = []
        for sheet_name in neighbor_sheets:
            device_name = sheet_name.replace("Neighbors_", "")
            device_names.append(device_name)
            logger.info(f"  {sheet_name} -> {device_name}")
        
        # Check for duplicates
        unique_devices = set(device_names)
        duplicates = []
        
        for device in unique_devices:
            count = device_names.count(device)
            if count > 1:
                duplicates.append((device, count))
        
        if duplicates:
            logger.error(f"\n❌ DUPLICATE SHEETS FOUND!")
            for device, count in duplicates:
                logger.error(f"  {device}: {count} sheets")
                
                # Show which sheets are duplicates
                matching_sheets = [s for s in neighbor_sheets if device in s]
                for sheet in matching_sheets:
                    logger.error(f"    - {sheet}")
                    
                    # Analyze the content of duplicate sheets
                    try:
                        ws = workbook[sheet]
                        header = ws['A1'].value if ws['A1'].value else "No header"
                        ip_info = ws['A2'].value if ws['A2'].value else "No IP info"
                        logger.error(f"      Header: {header}")
                        logger.error(f"      IP Info: {ip_info}")
                    except Exception as e:
                        logger.error(f"      Could not read sheet content: {e}")
        else:
            logger.info(f"✅ No duplicate sheets found")
        
        workbook.close()
        return len(duplicates) > 0
        
    except Exception as e:
        logger.error(f"Error analyzing {report_path}: {e}")
        return False

def main():
    """Main function to analyze existing seed reports"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("ANALYZING EXISTING SEED REPORTS FOR DUPLICATE SHEETS")
    logger.info("=" * 80)
    
    try:
        reports_dir = Path("reports")
        seed_reports = list(reports_dir.glob("Seed_*.xlsx"))
        
        if not seed_reports:
            logger.error("No seed reports found")
            return False
        
        logger.info(f"Found {len(seed_reports)} seed reports")
        
        # Analyze the most recent few reports
        recent_reports = sorted(seed_reports, key=lambda p: p.stat().st_mtime)[-5:]
        
        duplicates_found = False
        
        for report in recent_reports:
            logger.info(f"\n{'-' * 60}")
            has_duplicates = analyze_seed_report(report)
            if has_duplicates:
                duplicates_found = True
        
        logger.info(f"\n" + "=" * 80)
        logger.info("ANALYSIS SUMMARY")
        logger.info("=" * 80)
        
        if duplicates_found:
            logger.error("❌ DUPLICATE SHEETS ISSUE CONFIRMED")
            logger.error("Found duplicate sheets in seed workbooks")
            
            # Let's look at the Excel generator logic to understand why
            logger.info("\nInvestigating the root cause...")
            
            # The issue is likely in the device matching logic
            logger.info("Potential causes:")
            logger.info("1. Device hostname matching logic creates duplicates")
            logger.info("2. FQDN vs short hostname variations")
            logger.info("3. Case sensitivity issues")
            logger.info("4. Multiple IP addresses for same device")
            
            return False
        else:
            logger.info("✅ No duplicate sheets found in recent reports")
            return True
        
    except ImportError:
        logger.error("openpyxl not available")
        return False
    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)