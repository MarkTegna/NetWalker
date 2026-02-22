import openpyxl
import sys

# Open the latest Seed report
excel_file = 'marktests/reports/Seed_KARE-CORE-A_20260222-15-48.xlsx'

try:
    wb = openpyxl.load_workbook(excel_file)
    print(f"\n=== Sheets in {excel_file} ===\n")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    
    if 'Stack Members' in wb.sheetnames:
        ws = wb['Stack Members']
        print(f"\n=== Stack Members Sheet ===\n")
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
        
        # Print headers
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, col).value)
        print(f"\nHeaders: {headers}")
        
        # Print all data rows
        print(f"\nData:")
        for row in range(2, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                row_data.append(ws.cell(row, col).value)
            print(f"  Row {row}: {row_data}")
    else:
        print("\n!!! Stack Members sheet NOT FOUND !!!")
    
    wb.close()
except Exception as e:
    print(f"Error: {e}")
