#!/usr/bin/env python3
"""
Search database for 'ep-moldham' (case insensitive) across all tables and export to Excel
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Add the project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from netwalker.config.config_manager import ConfigurationManager
from netwalker.database import DatabaseManager


def search_database(db_manager, search_term):
    """
    Search all tables for the search term (case insensitive)

    Returns:
        Dictionary with table_name -> list of matching rows
    """
    results = {}

    cursor = db_manager.connection.cursor()

    # Get all tables in the database
    cursor.execute("""
        SELECT TABLE_NAME
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_TYPE = 'BASE TABLE' AND TABLE_SCHEMA = 'dbo'
        ORDER BY TABLE_NAME
    """)

    tables = [row[0] for row in cursor.fetchall()]

    print(f"\nSearching {len(tables)} tables for '{search_term}' (case insensitive)...")
    print("=" * 80)

    for table in tables:
        # Get column names for this table
        cursor.execute(f"""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = '{table}'
            ORDER BY ORDINAL_POSITION
        """)

        columns_info = cursor.fetchall()
        columns = [col[0] for col in columns_info]

        # Build WHERE clause to search all text columns
        text_columns = [col[0] for col in columns_info
                       if col[1] in ('nvarchar', 'varchar', 'nchar', 'char', 'text', 'ntext')]

        if not text_columns:
            continue

        # Build query with LIKE for each text column
        where_clauses = [f"{col} LIKE '%{search_term}%'" for col in text_columns]
        where_clause = " OR ".join(where_clauses)

        query = f"SELECT * FROM {table} WHERE {where_clause}"

        try:
            cursor.execute(query)
            rows = cursor.fetchall()

            if rows:
                print(f"\n[FOUND] {table}: {len(rows)} matching rows")
                results[table] = {
                    'columns': columns,
                    'rows': rows
                }
        except Exception as e:
            print(f"[ERROR] {table}: {e}")

    cursor.close()
    return results


def export_to_excel(results, output_file):
    """
    Export search results to Excel with one sheet per table
    """
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Create summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.append(["Search Results Summary"])
    summary_ws.append(["Search Term:", "ep-moldham (case insensitive)"])
    summary_ws.append(["Search Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append([])
    summary_ws.append(["Table Name", "Matching Rows"])

    # Style summary header
    for cell in summary_ws[1]:
        cell.font = Font(bold=True, size=14)

    for cell in summary_ws[5]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add data to summary
    total_matches = 0
    for table_name, data in sorted(results.items()):
        row_count = len(data['rows'])
        total_matches += row_count
        summary_ws.append([table_name, row_count])

    summary_ws.append([])
    summary_ws.append(["Total Matches:", total_matches])
    summary_ws["A" + str(summary_ws.max_row)].font = Font(bold=True)
    summary_ws["B" + str(summary_ws.max_row)].font = Font(bold=True)

    # Auto-size columns in summary
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width

    # Create a sheet for each table with results
    for table_name, data in sorted(results.items()):
        # Truncate sheet name if needed (Excel limit is 31 chars)
        sheet_name = table_name[:31] if len(table_name) > 31 else table_name

        ws = wb.create_sheet(sheet_name)

        # Add headers
        ws.append(data['columns'])

        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows
        for row in data['rows']:
            # Convert row to list and handle None values
            row_data = [str(val) if val is not None else '' for val in row]
            ws.append(row_data)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_file)
    print(f"\n[OK] Results exported to: {output_file}")


def main():
    """Main function"""

    # Load configuration
    config_file = 'netwalker.ini'
    config_manager = ConfigurationManager(config_file)
    parsed_config = config_manager.load_configuration()
    db_config = parsed_config.get('database', {})

    # Initialize database manager
    db_manager = DatabaseManager(db_config)

    if not db_manager.connect():
        print("[FAIL] Could not connect to database")
        return 1

    try:
        print("=" * 80)
        print("SEARCHING DATABASE FOR 'ep-moldham' (case insensitive)")
        print("=" * 80)

        # Search database
        results = search_database(db_manager, 'ep-moldham')

        if not results:
            print("\n[INFO] No matches found for 'ep-moldham'")
            return 0

        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        output_file = f"ep_moldham_search_results_{timestamp}.xlsx"

        # Export to Excel
        print("\n" + "=" * 80)
        print("EXPORTING RESULTS TO EXCEL")
        print("=" * 80)
        export_to_excel(results, output_file)

        # Print summary
        print("\n" + "=" * 80)
        print("SEARCH SUMMARY")
        print("=" * 80)
        total_matches = sum(len(data['rows']) for data in results.values())
        print(f"Tables with matches: {len(results)}")
        print(f"Total matching rows: {total_matches}")
        print("\nTables:")
        for table_name, data in sorted(results.items()):
            print(f"  - {table_name}: {len(data['rows'])} rows")
        print("=" * 80)

    except Exception as e:
        print(f"\n[FAIL] Error: {e}")
        import traceback
        traceback.print_exc()
        return 1

    finally:
        db_manager.disconnect()

    return 0


if __name__ == "__main__":
    sys.exit(main())
