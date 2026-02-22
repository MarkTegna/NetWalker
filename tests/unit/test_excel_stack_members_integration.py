"""
Integration tests for Stack Members sheet in inventory reports
Feature: switch-stack-member-collection
"""

import pytest
import os
from datetime import datetime
from openpyxl import load_workbook

from netwalker.reports.excel_generator import ExcelReportGenerator
from netwalker.connection.data_models import StackMemberInfo


class TestStackMembersIntegration:
    """Integration tests for Stack Members sheet in full report generation"""
    
    def setup_method(self):
        """Set up test fixtures"""
        self.config = {
            'reports_directory': './test_reports',
            'output': {
                'site_boundary_pattern': '*-CORE-*'
            }
        }
        self.generator = ExcelReportGenerator(self.config)
        
        # Ensure test reports directory exists
        os.makedirs('./test_reports', exist_ok=True)
    
    def teardown_method(self):
        """Clean up test files"""
        # Remove test report files
        if os.path.exists('./test_reports'):
            for file in os.listdir('./test_reports'):
                if file.startswith('Inventory_') and file.endswith('.xlsx'):
                    try:
                        os.remove(os.path.join('./test_reports', file))
                    except:
                        pass
    
    def test_stack_members_sheet_in_inventory_report(self):
        """
        Test that Stack Members sheet is created in inventory report
        
        Validates: Requirements 8.5 (Property 5: Sheet Creation Consistency)
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address="00:1a:2b:3c:4d:5e",
                software_version="16.9.1",
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address="00:1a:2b:3c:4d:5f",
                software_version="16.9.1",
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'software_version': '16.9.1',
                'vtp_version': '2',
                'serial_number': 'FCW1234A001',
                'hardware_model': 'WS-C3850-48P',
                'uptime': '1 day',
                'discovery_depth': 0,
                'discovery_method': 'seed',
                'discovery_timestamp': datetime.now(),
                'connection_method': 'SSH',
                'status': 'connected',
                'is_stack': True,
                'stack_members': stack_members,
                'vlans': []
            }
        }
        
        # Generate inventory report
        report_path = self.generator.generate_inventory_report(inventory)
        
        # Verify report was created
        assert os.path.exists(report_path), "Report file should be created"
        
        # Load workbook and verify Stack Members sheet exists
        workbook = load_workbook(report_path)
        
        assert "Stack Members" in workbook.sheetnames, "Stack Members sheet should exist in report"
        
        # Verify sheet ordering (Stack Members should come after Device Inventory)
        sheet_names = workbook.sheetnames
        device_inventory_index = sheet_names.index("Device Inventory")
        stack_members_index = sheet_names.index("Stack Members")
        
        assert stack_members_index > device_inventory_index, \
            "Stack Members sheet should come after Device Inventory sheet"
        
        workbook.close()
    
    def test_sheet_ordering_in_report(self):
        """
        Test that Stack Members sheet appears in correct position
        
        Validates: Requirements 8.5
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'software_version': '16.9.1',
                'vtp_version': '2',
                'serial_number': 'FCW1234A001',
                'hardware_model': 'WS-C3850-48P',
                'uptime': '1 day',
                'discovery_depth': 0,
                'discovery_method': 'seed',
                'discovery_timestamp': datetime.now(),
                'connection_method': 'SSH',
                'status': 'connected',
                'is_stack': True,
                'stack_members': stack_members,
                'vlans': []
            }
        }
        
        # Generate inventory report
        report_path = self.generator.generate_inventory_report(inventory)
        
        # Load workbook and verify sheet order
        workbook = load_workbook(report_path)
        sheet_names = workbook.sheetnames
        
        # Expected order: Device Inventory, Stack Members, VLAN Inventory
        assert "Device Inventory" in sheet_names
        assert "Stack Members" in sheet_names
        assert "VLAN Inventory" in sheet_names
        
        device_idx = sheet_names.index("Device Inventory")
        stack_idx = sheet_names.index("Stack Members")
        vlan_idx = sheet_names.index("VLAN Inventory")
        
        assert device_idx < stack_idx < vlan_idx, \
            "Sheets should be in order: Device Inventory, Stack Members, VLAN Inventory"
        
        workbook.close()
    
    def test_mixed_inventory_report(self):
        """
        Test report with both stacked and non-stacked devices
        
        Validates: Requirements 8.1, 8.4
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "StackDevice:192.168.1.1": {
                'hostname': 'StackDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'software_version': '16.9.1',
                'vtp_version': '2',
                'serial_number': 'FCW1234A001',
                'hardware_model': 'WS-C3850-48P',
                'uptime': '1 day',
                'discovery_depth': 0,
                'discovery_method': 'seed',
                'discovery_timestamp': datetime.now(),
                'connection_method': 'SSH',
                'status': 'connected',
                'is_stack': True,
                'stack_members': stack_members,
                'vlans': []
            },
            "StandaloneDevice:192.168.1.2": {
                'hostname': 'StandaloneDevice',
                'primary_ip': '192.168.1.2',
                'platform': 'cisco_ios',
                'software_version': '15.2.4',
                'vtp_version': '2',
                'serial_number': 'FDO1234A001',
                'hardware_model': 'WS-C2960X-48FPS',
                'uptime': '2 days',
                'discovery_depth': 1,
                'discovery_method': 'neighbor',
                'discovery_timestamp': datetime.now(),
                'connection_method': 'SSH',
                'status': 'connected',
                'is_stack': False,
                'stack_members': [],
                'vlans': []
            }
        }
        
        # Generate inventory report
        report_path = self.generator.generate_inventory_report(inventory)
        
        # Load workbook and verify Stack Members sheet
        workbook = load_workbook(report_path)
        ws = workbook["Stack Members"]
        
        # Should only have 1 data row (from StackDevice)
        assert ws.cell(row=2, column=1).value == "StackDevice"
        assert ws.cell(row=3, column=1).value is None  # No second row
        
        workbook.close()
    
    def test_empty_inventory_report(self):
        """
        Test report generation with empty inventory
        
        Validates: Requirements 8.4 (Property 6: Empty Sheet Handling)
        """
        inventory = {}
        
        # Generate inventory report
        report_path = self.generator.generate_inventory_report(inventory)
        
        # Load workbook and verify Stack Members sheet exists with headers only
        workbook = load_workbook(report_path)
        
        assert "Stack Members" in workbook.sheetnames, "Stack Members sheet should exist even with empty inventory"
        
        ws = workbook["Stack Members"]
        
        # Verify headers are present
        assert ws.cell(row=1, column=1).value == "Stack Device"
        
        # Verify no data rows
        assert ws.cell(row=2, column=1).value is None
        
        workbook.close()
