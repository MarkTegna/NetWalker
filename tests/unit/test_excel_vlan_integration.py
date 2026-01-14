"""
Unit tests for Excel Generator VLAN integration
Feature: vlan-inventory-collection
"""

import pytest
from unittest.mock import Mock, patch
from datetime import datetime
from openpyxl import Workbook

from netwalker.reports.excel_generator import ExcelReportGenerator
from netwalker.connection.data_models import VLANInfo


class TestExcelVLANIntegration:
    """Unit tests for Excel Generator VLAN integration"""
    
    def setup_method(self):
        """Set up test fixtures"""
        self.config = {
            'reports_directory': './test_reports',
            'output': {
                'site_boundary_pattern': '*-CORE-*'
            }
        }
        self.generator = ExcelReportGenerator(self.config)
    
    def test_vlan_inventory_sheet_creation(self):
        """
        Test that "VLAN Inventory" sheet is created in workbooks
        
        Validates: Requirements 4.1
        """
        # Create test inventory with VLAN data
        test_vlans = [
            VLANInfo(
                vlan_id=1,
                vlan_name="default",
                port_count=5,
                portchannel_count=1,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime.now()
            ),
            VLANInfo(
                vlan_id=10,
                vlan_name="VLAN_10",
                port_count=3,
                portchannel_count=0,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime.now()
            )
        ]
        
        inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': test_vlans,
                'vlan_collection_status': 'success'
            }
        }
        
        # Create workbook and add VLAN inventory sheet
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        self.generator._create_vlan_inventory_sheet(workbook, inventory)
        
        # Verify VLAN Inventory sheet was created
        assert "VLAN Inventory" in workbook.sheetnames, "VLAN Inventory sheet should be created"
        
        # Verify sheet has correct structure
        ws = workbook["VLAN Inventory"]
        
        # Check headers
        expected_headers = [
            "Device Name", "VLAN ID", "VLAN Name", "Port Count", "PortChannel Count", 
            "Device IP", "Platform", "Collection Timestamp", "Collection Status"
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"Header mismatch at column {col}: expected '{expected_header}', got '{actual_header}'"
    
    def test_vlan_sheet_column_structure(self):
        """
        Test that required columns are present in correct order
        
        Validates: Requirements 4.2
        """
        inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': [],
                'vlan_collection_status': 'success'
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_vlan_inventory_sheet(workbook, inventory)
        ws = workbook["VLAN Inventory"]
        
        # Verify column structure
        expected_columns = [
            ("Device Name", 1),
            ("VLAN ID", 2), 
            ("VLAN Name", 3),
            ("Port Count", 4),
            ("PortChannel Count", 5),
            ("Device IP", 6),
            ("Platform", 7),
            ("Collection Timestamp", 8),
            ("Collection Status", 9)
        ]
        
        for header_name, col_num in expected_columns:
            actual_value = ws.cell(row=1, column=col_num).value
            assert actual_value == header_name, f"Column {col_num} should be '{header_name}', got '{actual_value}'"
    
    def test_vlan_data_row_organization(self):
        """
        Test VLAN data row organization property
        
        Validates: Requirements 4.3 (Property 14: VLAN Data Row Organization)
        """
        # Create test VLANs with specific data
        test_vlans = [
            VLANInfo(
                vlan_id=1,
                vlan_name="default",
                port_count=5,
                portchannel_count=1,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime(2024, 1, 1, 12, 0, 0)
            ),
            VLANInfo(
                vlan_id=10,
                vlan_name="VLAN_10",
                port_count=3,
                portchannel_count=0,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime(2024, 1, 1, 12, 0, 0)
            )
        ]
        
        inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': test_vlans,
                'vlan_collection_status': 'success'
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_vlan_inventory_sheet(workbook, inventory)
        ws = workbook["VLAN Inventory"]
        
        # Verify data rows
        # Row 2 should contain VLAN 1 data
        assert ws.cell(row=2, column=1).value == "TestDevice"  # Device Name
        assert ws.cell(row=2, column=2).value == 1  # VLAN ID
        assert ws.cell(row=2, column=3).value == "default"  # VLAN Name
        assert ws.cell(row=2, column=4).value == 5  # Port Count
        assert ws.cell(row=2, column=5).value == 1  # PortChannel Count
        assert ws.cell(row=2, column=6).value == "192.168.1.1"  # Device IP
        assert ws.cell(row=2, column=7).value == "IOS"  # Platform
        assert ws.cell(row=2, column=9).value == "success"  # Collection Status
        
        # Row 3 should contain VLAN 10 data
        assert ws.cell(row=3, column=1).value == "TestDevice"  # Device Name
        assert ws.cell(row=3, column=2).value == 10  # VLAN ID
        assert ws.cell(row=3, column=3).value == "VLAN_10"  # VLAN Name
        assert ws.cell(row=3, column=4).value == 3  # Port Count
        assert ws.cell(row=3, column=5).value == 0  # PortChannel Count
    
    def test_professional_formatting_consistency(self):
        """
        Test professional formatting consistency property
        
        Validates: Requirements 4.4 (Property 15: Professional Formatting Consistency)
        """
        test_vlans = [
            VLANInfo(
                vlan_id=1,
                vlan_name="default",
                port_count=5,
                portchannel_count=1,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime.now()
            )
        ]
        
        inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': test_vlans,
                'vlan_collection_status': 'success'
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_vlan_inventory_sheet(workbook, inventory)
        ws = workbook["VLAN Inventory"]
        
        # Check header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold == True, "Header should be bold"
        assert header_cell.fill.start_color.rgb == "FF366092", "Header should have correct background color"
        assert header_cell.alignment.horizontal == "center", "Header should be center-aligned"
        
        # Check table creation
        assert len(ws.tables) == 1, "Should have exactly one table"
        table = list(ws.tables.values())[0]
        assert table.displayName == "VLANInventory", "Table should have correct name"
    
    def test_empty_vlan_data_handling(self):
        """
        Test empty sheet creation with appropriate headers and message
        
        Validates: Requirements 4.5
        """
        # Create inventory with no VLAN data
        inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': [],
                'vlan_collection_status': 'no_vlans_found'
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_vlan_inventory_sheet(workbook, inventory)
        ws = workbook["VLAN Inventory"]
        
        # Verify headers are still present
        expected_headers = [
            "Device Name", "VLAN ID", "VLAN Name", "Port Count", "PortChannel Count", 
            "Device IP", "Platform", "Collection Timestamp", "Collection Status"
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"Header should be present even with no data"
        
        # Verify appropriate message for empty data
        assert ws.cell(row=2, column=1).value == "No VLAN data available", "Should show appropriate message for empty data"
        assert "VLAN collection may be disabled" in str(ws.cell(row=2, column=2).value), "Should explain why no data is available"
    
    def test_consolidate_vlan_data_functionality(self):
        """
        Test the _consolidate_vlan_data helper method
        """
        # Create test inventory with multiple devices and VLANs
        test_vlans_device1 = [
            VLANInfo(
                vlan_id=1,
                vlan_name="default",
                port_count=5,
                portchannel_count=1,
                device_hostname="Device1",
                device_ip="192.168.1.1",
                collection_timestamp=datetime.now()
            )
        ]
        
        test_vlans_device2 = [
            VLANInfo(
                vlan_id=10,
                vlan_name="VLAN_10",
                port_count=3,
                portchannel_count=0,
                device_hostname="Device2",
                device_ip="192.168.1.2",
                collection_timestamp=datetime.now()
            )
        ]
        
        inventory = {
            "Device1:192.168.1.1": {
                'hostname': 'Device1',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': test_vlans_device1,
                'vlan_collection_status': 'success'
            },
            "Device2:192.168.1.2": {
                'hostname': 'Device2',
                'primary_ip': '192.168.1.2',
                'platform': 'NX-OS',
                'vlans': test_vlans_device2,
                'vlan_collection_status': 'success'
            }
        }
        
        # Test consolidation
        vlan_data = self.generator._consolidate_vlan_data(inventory)
        
        # Verify consolidated data
        assert len(vlan_data) == 2, "Should consolidate VLANs from both devices"
        
        # Check first VLAN entry
        vlan1 = vlan_data[0]
        assert vlan1['device_name'] == 'Device1'
        assert vlan1['vlan_id'] == 1
        assert vlan1['vlan_name'] == 'default'
        assert vlan1['port_count'] == 5
        assert vlan1['portchannel_count'] == 1
        
        # Check second VLAN entry
        vlan2 = vlan_data[1]
        assert vlan2['device_name'] == 'Device2'
        assert vlan2['vlan_id'] == 10
        assert vlan2['vlan_name'] == 'VLAN_10'
        assert vlan2['port_count'] == 3
        assert vlan2['portchannel_count'] == 0
    
    def test_master_vlan_inventory_sheet_creation(self):
        """
        Test master VLAN inventory sheet creation for multi-seed discovery
        """
        # Create consolidated inventory with seed information
        test_vlans = [
            VLANInfo(
                vlan_id=1,
                vlan_name="default",
                port_count=5,
                portchannel_count=1,
                device_hostname="TestDevice",
                device_ip="192.168.1.1",
                collection_timestamp=datetime.now()
            )
        ]
        
        consolidated_inventory = {
            "TestDevice:192.168.1.1": {
                'hostname': 'TestDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'IOS',
                'vlans': test_vlans,
                'vlan_collection_status': 'success',
                'discovered_by_seeds': ['seed1', 'seed2']
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_master_vlan_inventory_sheet(workbook, consolidated_inventory)
        
        # Verify master VLAN inventory sheet was created
        assert "Master VLAN Inventory" in workbook.sheetnames, "Master VLAN Inventory sheet should be created"
        
        ws = workbook["Master VLAN Inventory"]
        
        # Check that seed information is included
        expected_headers = [
            "Device Name", "VLAN ID", "VLAN Name", "Port Count", "PortChannel Count", 
            "Device IP", "Platform", "Collection Timestamp", "Discovered By Seeds"
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"Master sheet header mismatch at column {col}"
        
        # Verify seed information is populated
        assert ws.cell(row=2, column=9).value == "seed1, seed2", "Should include seed discovery information"