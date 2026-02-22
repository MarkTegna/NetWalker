"""
Unit tests for Excel Generator Stack Members sheet
Feature: switch-stack-member-collection
"""

import pytest
from datetime import datetime
from openpyxl import Workbook

from netwalker.reports.excel_generator import ExcelReportGenerator
from netwalker.connection.data_models import StackMemberInfo


class TestExcelStackMembersSheet:
    """Unit tests for Stack Members sheet creation"""
    
    def setup_method(self):
        """Set up test fixtures"""
        self.config = {
            'reports_directory': './test_reports',
            'output': {
                'site_boundary_pattern': '*-CORE-*'
            }
        }
        self.generator = ExcelReportGenerator(self.config)
    
    def test_stack_members_sheet_creation(self):
        """
        Test that Stack Members sheet is created with proper headers
        
        Validates: Requirements 8.1, 8.5
        """
        # Create test inventory with stack members
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address="00:1a:2b:3c:4d:5e",
                software_version="16.9.1",
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address="00:1a:2b:3c:4d:5f",
                software_version="16.9.1",
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        # Create workbook and add stack members sheet
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        
        # Verify Stack Members sheet was created
        assert "Stack Members" in workbook.sheetnames, "Stack Members sheet should be created"
        
        # Verify sheet has correct structure
        ws = workbook["Stack Members"]
        
        # Check headers
        expected_headers = [
            "Stack Device", "Stack IP", "Platform", "Switch Number", "Role",
            "Priority", "Hardware Model", "Serial Number", "MAC Address",
            "Software Version", "State"
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, \
                f"Header mismatch at column {col}: expected '{expected_header}', got '{actual_header}'"
    
    def test_single_stack_data_rows(self):
        """
        Test that stack member data is written correctly for a single stack
        
        Validates: Requirements 8.2
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address="00:1a:2b:3c:4d:5e",
                software_version="16.9.1",
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address="00:1a:2b:3c:4d:5f",
                software_version="16.9.1",
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify first stack member row
        assert ws.cell(row=2, column=1).value == "TestStack"
        assert ws.cell(row=2, column=2).value == "192.168.1.1"
        assert ws.cell(row=2, column=3).value == "cisco_ios"
        assert ws.cell(row=2, column=4).value == 1
        assert ws.cell(row=2, column=5).value == "Master"
        assert ws.cell(row=2, column=6).value == 15
        assert ws.cell(row=2, column=7).value == "WS-C3850-48P"
        assert ws.cell(row=2, column=8).value == "FCW1234A001"
        assert ws.cell(row=2, column=9).value == "00:1a:2b:3c:4d:5e"
        assert ws.cell(row=2, column=10).value == "16.9.1"
        assert ws.cell(row=2, column=11).value == "Ready"
        
        # Verify second stack member row
        assert ws.cell(row=3, column=1).value == "TestStack"
        assert ws.cell(row=3, column=2).value == "192.168.1.1"
        assert ws.cell(row=3, column=4).value == 2
        assert ws.cell(row=3, column=5).value == "Member"
        assert ws.cell(row=3, column=8).value == "FCW1234A002"
    
    def test_multiple_stacks_data_rows(self):
        """
        Test that multiple stacks are handled correctly
        
        Validates: Requirements 8.1, 8.2
        """
        stack1_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        stack2_members = [
            StackMemberInfo(
                switch_number=1,
                role="Active",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE1234A001",
                mac_address=None,
                software_version=None,
                state="Ok"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Standby",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE1234A002",
                mac_address=None,
                software_version=None,
                state="Ok"
            )
        ]
        
        inventory = {
            "Stack1:192.168.1.1": {
                'hostname': 'Stack1',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack1_members
            },
            "Stack2:192.168.1.2": {
                'hostname': 'Stack2',
                'primary_ip': '192.168.1.2',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack2_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Should have 3 data rows (1 from Stack1, 2 from Stack2)
        # Verify we have the correct number of rows
        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=3, column=1).value is not None
        assert ws.cell(row=4, column=1).value is not None
        assert ws.cell(row=5, column=1).value is None  # No 5th row
    
    def test_non_stack_devices_excluded(self):
        """
        Test that non-stack devices do not contribute rows
        
        Validates: Requirements 8.4 (Property 2: Non-Stack Device Exclusion)
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "StackDevice:192.168.1.1": {
                'hostname': 'StackDevice',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            },
            "StandaloneDevice:192.168.1.2": {
                'hostname': 'StandaloneDevice',
                'primary_ip': '192.168.1.2',
                'platform': 'cisco_ios',
                'is_stack': False,
                'stack_members': []
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Should only have 1 data row (from StackDevice)
        assert ws.cell(row=2, column=1).value == "StackDevice"
        assert ws.cell(row=3, column=1).value is None  # No second row
    
    def test_missing_optional_fields(self):
        """
        Test that stack members with missing optional fields are handled correctly
        
        Validates: Requirements 8.2 (Property 3: Required Field Presence)
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=None,  # Optional field missing
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,  # Optional field missing
                software_version=None,  # Optional field missing
                state=None  # Optional field missing
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify row is created with empty cells for optional fields
        assert ws.cell(row=2, column=1).value == "TestStack"
        assert ws.cell(row=2, column=4).value == 1  # Switch number (required)
        assert ws.cell(row=2, column=5).value == "Master"  # Role
        assert ws.cell(row=2, column=6).value == ""  # Priority (optional, None)
        assert ws.cell(row=2, column=7).value == "WS-C3850-48P"  # Hardware model (required)
        assert ws.cell(row=2, column=8).value == "FCW1234A001"  # Serial number (required)
        assert ws.cell(row=2, column=9).value == ""  # MAC address (optional, None)
        assert ws.cell(row=2, column=10).value == ""  # Software version (optional, None)
        assert ws.cell(row=2, column=11).value == ""  # State (optional, None)
    
    def test_missing_required_fields_skipped(self):
        """
        Test that stack members with missing required fields are skipped with warning
        
        Validates: Requirements 8.2
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=None,  # Missing required field
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Should only have 1 data row (second member skipped due to missing switch_number)
        assert ws.cell(row=2, column=1).value == "TestStack"
        assert ws.cell(row=2, column=4).value == 1
        assert ws.cell(row=3, column=1).value is None  # No second row
    
    def test_empty_inventory(self):
        """
        Test that empty inventory creates sheet with headers only
        
        Validates: Requirements 8.4 (Property 6: Empty Sheet Handling)
        """
        inventory = {}
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify headers are present
        assert ws.cell(row=1, column=1).value == "Stack Device"
        assert ws.cell(row=1, column=2).value == "Stack IP"
        
        # Verify no data rows
        assert ws.cell(row=2, column=1).value is None
    
    def test_no_stacks_in_inventory(self):
        """
        Test that inventory with no stacks creates sheet with headers only
        
        Validates: Requirements 8.4 (Property 6: Empty Sheet Handling)
        """
        inventory = {
            "Device1:192.168.1.1": {
                'hostname': 'Device1',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': False,
                'stack_members': []
            },
            "Device2:192.168.1.2": {
                'hostname': 'Device2',
                'primary_ip': '192.168.1.2',
                'platform': 'cisco_ios',
                'is_stack': False,
                'stack_members': []
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify headers are present
        assert ws.cell(row=1, column=1).value == "Stack Device"
        
        # Verify no data rows
        assert ws.cell(row=2, column=1).value is None
    
    def test_stack_role_identification(self):
        """
        Test that master/active roles are clearly displayed
        
        Validates: Requirements 8.3 (Property 4: Stack Role Identification)
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify master role is clearly displayed
        assert ws.cell(row=2, column=5).value == "Master"
        assert ws.cell(row=3, column=5).value == "Member"
    
    def test_vss_stack_configuration(self):
        """
        Test VSS (Virtual Switching System) stack configuration
        
        Validates: Requirements 8.2, 8.3
        """
        vss_members = [
            StackMemberInfo(
                switch_number=1,
                role="Active",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE1234A001",
                mac_address=None,
                software_version=None,
                state="Ok"
            ),
            StackMemberInfo(
                switch_number=2,
                role="Standby",
                priority=None,
                hardware_model="WS-C4500X-32",
                serial_number="JAE1234A002",
                mac_address=None,
                software_version=None,
                state="Ok"
            )
        ]
        
        inventory = {
            "VSSStack:192.168.1.1": {
                'hostname': 'VSSStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': vss_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify VSS roles are displayed correctly
        assert ws.cell(row=2, column=5).value == "Active"
        assert ws.cell(row=3, column=5).value == "Standby"
        assert ws.cell(row=2, column=7).value == "WS-C4500X-32"
        assert ws.cell(row=3, column=7).value == "WS-C4500X-32"
    
    def test_table_formatting_applied(self):
        """
        Test that Excel table formatting is applied correctly
        
        Validates: Requirements 8.5
        """
        stack_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "TestStack:192.168.1.1": {
                'hostname': 'TestStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify table was created
        assert len(ws.tables) == 1, "Should have exactly one table"
        table = list(ws.tables.values())[0]
        assert table.displayName == "StackMembers", "Table should have correct name"
        
        # Verify header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold == True, "Header should be bold"
    
    def test_data_sorting(self):
        """
        Test that data is sorted by stack device name, then switch number
        
        Validates: Requirements 8.2
        """
        stack1_members = [
            StackMemberInfo(
                switch_number=2,
                role="Member",
                priority=1,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A002",
                mac_address=None,
                software_version=None,
                state="Ready"
            ),
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234A001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        stack2_members = [
            StackMemberInfo(
                switch_number=1,
                role="Master",
                priority=15,
                hardware_model="WS-C3850-48P",
                serial_number="FCW1234B001",
                mac_address=None,
                software_version=None,
                state="Ready"
            )
        ]
        
        inventory = {
            "ZStack:192.168.1.2": {
                'hostname': 'ZStack',
                'primary_ip': '192.168.1.2',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack2_members
            },
            "AStack:192.168.1.1": {
                'hostname': 'AStack',
                'primary_ip': '192.168.1.1',
                'platform': 'cisco_ios',
                'is_stack': True,
                'stack_members': stack1_members
            }
        }
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        self.generator._create_stack_members_sheet(workbook, inventory)
        ws = workbook["Stack Members"]
        
        # Verify sorting: AStack should come before ZStack
        # Within AStack, switch 1 should come before switch 2
        assert ws.cell(row=2, column=1).value == "AStack"
        assert ws.cell(row=2, column=4).value == 1
        assert ws.cell(row=3, column=1).value == "AStack"
        assert ws.cell(row=3, column=4).value == 2
        assert ws.cell(row=4, column=1).value == "ZStack"
        assert ws.cell(row=4, column=4).value == 1
