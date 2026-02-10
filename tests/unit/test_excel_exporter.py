"""
Unit Tests for Excel Exporter

Tests the CommandResultExporter class for exporting command results to Excel.

Author: Mark Oldham
"""

import os
import pytest
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from netwalker.executor.excel_exporter import CommandResultExporter
from netwalker.executor.data_models import CommandResult


class TestCommandResultExporter:
    """Test suite for CommandResultExporter class"""
    
    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """Create a temporary output directory for tests"""
        output_dir = tmp_path / "excel_output"
        output_dir.mkdir()
        return str(output_dir)
    
    @pytest.fixture
    def exporter(self, temp_output_dir):
        """Create a CommandResultExporter instance"""
        return CommandResultExporter(output_dir=temp_output_dir)
    
    @pytest.fixture
    def sample_results(self):
        """Create sample command results for testing"""
        return [
            CommandResult(
                device_name="ROUTER-01",
                ip_address="10.1.1.1",
                status="Success",
                output="Interface GigabitEthernet0/0\n  IP Address: 10.1.1.1\n  Status: up",
                execution_time=1.23
            ),
            CommandResult(
                device_name="SWITCH-02",
                ip_address="10.1.1.2",
                status="Failed",
                output="Connection timeout",
                execution_time=30.0
            ),
            CommandResult(
                device_name="ROUTER-03",
                ip_address="10.1.1.3",
                status="Success",
                output="show version\nCisco IOS Software, Version 15.2",
                execution_time=2.45
            )
        ]
    
    def test_exporter_initialization(self, temp_output_dir):
        """Test that exporter initializes correctly"""
        exporter = CommandResultExporter(output_dir=temp_output_dir)
        assert exporter.output_dir == temp_output_dir
        assert Path(temp_output_dir).exists()
    
    def test_exporter_creates_output_directory(self, tmp_path):
        """Test that exporter creates output directory if it doesn't exist"""
        output_dir = tmp_path / "new_directory"
        assert not output_dir.exists()
        
        exporter = CommandResultExporter(output_dir=str(output_dir))
        assert output_dir.exists()
    
    def test_export_creates_file(self, exporter, sample_results, temp_output_dir):
        """Test that export creates an Excel file"""
        filepath = exporter.export(sample_results)
        
        assert os.path.exists(filepath)
        assert filepath.startswith(temp_output_dir)
        assert filepath.endswith(".xlsx")
    
    def test_export_filename_format(self, exporter, sample_results):
        """Test that export uses correct filename format: Command_Results_YYYYMMDD-HH-MM.xlsx"""
        filepath = exporter.export(sample_results)
        filename = os.path.basename(filepath)
        
        # Check filename pattern
        assert filename.startswith("Command_Results_")
        assert filename.endswith(".xlsx")
        
        # Extract timestamp part
        timestamp_part = filename.replace("Command_Results_", "").replace(".xlsx", "")
        
        # Verify timestamp format (YYYYMMDD-HH-MM)
        try:
            datetime.strptime(timestamp_part, "%Y%m%d-%H-%M")
        except ValueError:
            pytest.fail(f"Filename timestamp '{timestamp_part}' does not match format YYYYMMDD-HH-MM")
    
    def test_export_worksheet_structure(self, exporter, sample_results):
        """Test that exported Excel has correct worksheet structure"""
        filepath = exporter.export(sample_results)
        
        # Load workbook and check structure
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check worksheet title
        assert ws.title == "Command Results"
        
        # Check headers
        expected_headers = ["Device Name", "Device IP", "Status", "Command Output", "Execution Time"]
        for col_num, expected_header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_num).value == expected_header
        
        wb.close()
    
    def test_export_header_formatting(self, exporter, sample_results):
        """Test that headers have correct formatting (bold white on blue #366092)"""
        filepath = exporter.export(sample_results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check first header cell formatting
        header_cell = ws.cell(row=1, column=1)
        
        # Check font (bold, white color)
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb == "FFFFFFFF" or header_cell.font.color.rgb == "00FFFFFF"
        
        # Check fill (blue #366092)
        assert header_cell.fill.start_color.rgb == "FF366092" or header_cell.fill.start_color.rgb == "00366092"
        
        # Check alignment (center)
        assert header_cell.alignment.horizontal == "center"
        assert header_cell.alignment.vertical == "center"
        
        wb.close()
    
    def test_export_data_content(self, exporter, sample_results):
        """Test that exported data matches input results"""
        filepath = exporter.export(sample_results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check data rows (starting from row 2)
        for row_num, result in enumerate(sample_results, start=2):
            assert ws.cell(row=row_num, column=1).value == result.device_name
            assert ws.cell(row=row_num, column=2).value == result.ip_address
            assert ws.cell(row=row_num, column=3).value == result.status
            assert ws.cell(row=row_num, column=4).value == result.output
            assert ws.cell(row=row_num, column=5).value == f"{result.execution_time:.2f}s"
        
        wb.close()
    
    def test_export_preserves_line_breaks(self, exporter):
        """Test that line breaks in command output are preserved with wrap text"""
        results = [
            CommandResult(
                device_name="TEST-DEVICE",
                ip_address="10.0.0.1",
                status="Success",
                output="Line 1\nLine 2\nLine 3\nLine 4",
                execution_time=1.0
            )
        ]
        
        filepath = exporter.export(results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check that output cell has wrap_text enabled
        output_cell = ws.cell(row=2, column=4)
        assert output_cell.value == "Line 1\nLine 2\nLine 3\nLine 4"
        assert output_cell.alignment.wrap_text is True
        assert output_cell.alignment.vertical == "top"
        
        wb.close()
    
    def test_export_column_width_adjustment(self, exporter, sample_results):
        """Test that column widths are auto-adjusted"""
        filepath = exporter.export(sample_results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check that column widths are set (not default)
        for col_num in range(1, 6):
            column_letter = chr(64 + col_num)  # A, B, C, D, E
            width = ws.column_dimensions[column_letter].width
            assert width > 0, f"Column {column_letter} width should be set"
        
        wb.close()
    
    def test_export_column_width_max_limit(self, exporter):
        """Test that column widths are capped at 100 characters"""
        # Create result with very long output
        long_output = "A" * 200  # 200 character string
        results = [
            CommandResult(
                device_name="TEST",
                ip_address="10.0.0.1",
                status="Success",
                output=long_output,
                execution_time=1.0
            )
        ]
        
        filepath = exporter.export(results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check that Command Output column (D) is capped at 100
        width = ws.column_dimensions['D'].width
        assert width <= 100, f"Column width {width} should be capped at 100"
        
        wb.close()
    
    def test_export_empty_results(self, exporter):
        """Test exporting empty results list"""
        filepath = exporter.export([])
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Should have headers but no data rows
        assert ws.cell(row=1, column=1).value == "Device Name"
        assert ws.cell(row=2, column=1).value is None
        
        wb.close()
    
    def test_export_with_command_parameter(self, exporter, sample_results):
        """Test export with command parameter (for reference)"""
        command = "show ip eigrp neighbors"
        filepath = exporter.export(sample_results, command=command)
        
        # Should still create file successfully
        assert os.path.exists(filepath)
    
    def test_export_execution_time_formatting(self, exporter):
        """Test that execution time is formatted to 2 decimal places"""
        results = [
            CommandResult(
                device_name="TEST",
                ip_address="10.0.0.1",
                status="Success",
                output="output",
                execution_time=1.23456789
            )
        ]
        
        filepath = exporter.export(results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check execution time formatting
        time_cell = ws.cell(row=2, column=5)
        assert time_cell.value == "1.23s"
        
        wb.close()
    
    def test_export_handles_multiline_column_width(self, exporter):
        """Test that column width calculation handles multi-line content correctly"""
        results = [
            CommandResult(
                device_name="TEST",
                ip_address="10.0.0.1",
                status="Success",
                output="Short line\nThis is a much longer line that should determine the width\nShort",
                execution_time=1.0
            )
        ]
        
        filepath = exporter.export(results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Column width should be based on longest line, not total length
        width = ws.column_dimensions['D'].width
        # The longest line is about 60 chars, so width should be around that (plus padding)
        assert 50 < width < 100
        
        wb.close()
    
    def test_export_different_status_values(self, exporter):
        """Test export with different status values"""
        results = [
            CommandResult("DEV1", "10.0.0.1", "Success", "output", 1.0),
            CommandResult("DEV2", "10.0.0.2", "Failed", "error", 2.0),
            CommandResult("DEV3", "10.0.0.3", "Timeout", "timeout error", 30.0),
            CommandResult("DEV4", "10.0.0.4", "Auth Failed", "auth error", 5.0),
        ]
        
        filepath = exporter.export(results)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Verify all status values are recorded
        assert ws.cell(row=2, column=3).value == "Success"
        assert ws.cell(row=3, column=3).value == "Failed"
        assert ws.cell(row=4, column=3).value == "Timeout"
        assert ws.cell(row=5, column=3).value == "Auth Failed"
        
        wb.close()
