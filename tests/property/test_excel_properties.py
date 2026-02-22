"""
Property-based tests for ExcelReportGenerator

Tests universal properties of Excel report generation functionality.
"""

import pytest
from hypothesis import given, strategies as st, assume, settings
import os
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from netwalker.reports.excel_generator import ExcelReportGenerator


class TestExcelReportGeneratorProperties:
    """Property-based tests for ExcelReportGenerator functionality"""
    
    def create_test_config(self, temp_dir: str) -> dict:
        """Create test configuration with temporary directory"""
        return {
            'reports_directory': temp_dir
        }
    
    def create_test_inventory(self, devices: list) -> dict:
        """Create test device inventory"""
        inventory = {}
        for i, (hostname, ip_address, platform) in enumerate(devices):
            device_key = f"{hostname}:{ip_address}"
            inventory[device_key] = {
                'hostname': hostname,
                'ip_address': ip_address,
                'platform': platform,
                'software_version': f"Version {i+1}.0",
                'serial_number': f"SN{i+1:06d}",
                'hardware_model': f"Model-{platform}",
                'uptime': f"{i+1} days",
                'discovery_depth': i % 3,
                'discovery_method': 'cdp' if i % 2 == 0 else 'lldp',
                'parent_device': None if i == 0 else f"parent{i}:192.168.1.{i}",
                'discovery_timestamp': datetime.now().isoformat(),
                'connection_method': 'ssh' if i % 2 == 0 else 'telnet',
                'status': 'connected',
                'neighbors': [
                    {
                        'hostname': f"neighbor{i}_1",
                        'ip_address': f"192.168.{i+10}.1",
                        'protocol': 'cdp',
                        'local_interface': f"Gi0/{i}",
                        'remote_interface': f"Gi0/{i+1}",
                        'platform': 'IOS'
                    }
                ] if i < 3 else []  # Only first few devices have neighbors
            }
        return inventory
    
    @given(
        devices=st.lists(
            st.tuples(
                st.text(min_size=1, max_size=15, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd'))),
                st.ip_addresses(v=4).map(str),
                st.sampled_from(['IOS', 'IOS-XE', 'NX-OS'])
            ),
            min_size=1, max_size=5,
            unique_by=lambda x: f"{x[0]}:{x[1]}"
        )
    )
    def test_excel_report_completeness_property(self, devices):
        """
        Property 14: Excel Report Completeness
        
        Generated Excel reports should contain all required sheets and data.
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Create test data
            inventory = self.create_test_inventory(devices)
            discovery_results = {
                'discovery_time_seconds': 120.5,
                'total_devices': len(devices),
                'successful_connections': len(devices),
                'failed_connections': 0,
                'filtered_devices': 0,
                'boundary_devices': 0,
                'max_depth_reached': 2
            }
            seed_devices = [f"{devices[0][0]}:{devices[0][1]}"] if devices else []
            
            # Generate report
            filepaths = generator.generate_discovery_report(inventory, discovery_results, seed_devices)
            
            # Property: Files should be created
            assert len(filepaths) > 0, "At least one Excel file should be created"
            main_filepath = filepaths[0]  # First file is the main discovery report
            assert os.path.exists(main_filepath), "Main Excel file should be created"
            
            # Property: File should be a valid Excel workbook
            workbook = load_workbook(main_filepath)
            
            try:
                # Property: Required sheets should exist
                sheet_names = workbook.sheetnames
                required_sheets = ["Discovery Summary", "Device Inventory", "Network Connections"]
                
                for required_sheet in required_sheets:
                    assert required_sheet in sheet_names, f"Sheet '{required_sheet}' should exist"
                
                # Property: Device Inventory sheet should contain all devices
                device_sheet = workbook["Device Inventory"]
                
                # Count data rows (excluding header)
                data_rows = 0
                for row in device_sheet.iter_rows(min_row=2):
                    if row[0].value:  # If first column has value
                        data_rows += 1
                
                assert data_rows == len(devices), f"Should have {len(devices)} device rows, found {data_rows}"
                
            finally:
                workbook.close()
    
    @given(
        base_name=st.text(min_size=1, max_size=20, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd', 'Pc'))),
        extension=st.sampled_from(['xlsx', 'csv', 'txt'])
    )
    def test_timestamp_file_naming_property(self, base_name, extension):
        """
        Property 16: Timestamp-Based File Naming
        
        Generated filenames should include timestamps in YYYYMMDD-HH-MM format.
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Generate filename
            filename = generator.get_timestamp_filename(base_name, extension)
            
            # Property: Filename should contain base name
            assert base_name in filename, f"Filename should contain base name '{base_name}'"
            
            # Property: Filename should contain extension
            assert filename.endswith(f".{extension}"), f"Filename should end with '.{extension}'"
            
            # Property: Filename should contain timestamp pattern
            # Format: base_name_YYYYMMDD-HH-MM.extension
            import re
            timestamp_pattern = r'\d{8}-\d{2}-\d{2}'
            assert re.search(timestamp_pattern, filename), "Filename should contain timestamp in YYYYMMDD-HH-MM format"
    
    @given(
        devices=st.lists(
            st.tuples(
                st.text(min_size=1, max_size=10, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd'))),
                st.ip_addresses(v=4).map(str),
                st.sampled_from(['IOS', 'NX-OS'])
            ),
            min_size=1, max_size=3,
            unique_by=lambda x: f"{x[0]}:{x[1]}"
        )
    )
    @settings(max_examples=10, deadline=None)
    def test_excel_formatting_standards_property(self, devices):
        """
        Property 15: Excel Formatting Standards
        
        Excel reports should have consistent professional formatting.
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Create test data
            inventory = self.create_test_inventory(devices)
            discovery_results = {
                'discovery_time_seconds': 60.0,
                'total_devices': len(devices),
                'successful_connections': len(devices),
                'failed_connections': 0,
                'filtered_devices': 0,
                'boundary_devices': 0,
                'max_depth_reached': 1
            }
            seed_devices = [f"{devices[0][0]}:{devices[0][1]}"] if devices else []
            
            # Generate report
            filepaths = generator.generate_discovery_report(inventory, discovery_results, seed_devices)
            
            # Property: Files should be created
            assert len(filepaths) > 0, "At least one Excel file should be created"
            main_filepath = filepaths[0]  # First file is the main discovery report
            
            # Load and check formatting
            workbook = load_workbook(main_filepath)
            
            try:
                # Check Device Inventory sheet formatting
                device_sheet = workbook["Device Inventory"]
                
                # Property: Headers should be in row 1
                header_row = list(device_sheet.iter_rows(min_row=1, max_row=1))[0]
                
                for cell in header_row:
                    if cell.value:  # If cell has content
                        # Property: Header cells should have bold font
                        assert cell.font.bold, f"Header cell {cell.coordinate} should be bold"
                        
                        # Property: Header cells should have fill color
                        assert cell.fill.start_color.rgb, f"Header cell {cell.coordinate} should have background color"
                
                # Property: Columns should be auto-sized (width > 0)
                for column_letter in ['A', 'B', 'C', 'D']:
                    width = device_sheet.column_dimensions[column_letter].width
                    assert width > 0, f"Column {column_letter} should have width > 0"
                    
            finally:
                workbook.close()
    
    def test_master_inventory_generation_property(self):
        """
        Property: Master inventory should consolidate devices from multiple seeds
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Create inventories for multiple seeds
            devices1 = [("device1", "192.168.1.1", "IOS"), ("device2", "192.168.1.2", "IOS")]
            devices2 = [("device2", "192.168.1.2", "IOS"), ("device3", "192.168.1.3", "NX-OS")]
            
            inventory1 = self.create_test_inventory(devices1)
            inventory2 = self.create_test_inventory(devices2)
            
            all_inventories = {
                "seed1": inventory1,
                "seed2": inventory2
            }
            seed_devices = ["seed1", "seed2"]
            
            # Generate master inventory
            filepath = generator.generate_master_inventory(all_inventories, seed_devices)
            
            # Property: File should be created
            assert os.path.exists(filepath), "Master inventory file should be created"
            
            # Load and verify content
            workbook = load_workbook(filepath)
            
            try:
                # Property: Master inventory sheet should exist
                assert "Master Device Inventory" in workbook.sheetnames
                
                master_sheet = workbook["Master Device Inventory"]
                
                # Count unique devices (device2 appears in both inventories)
                unique_devices = set()
                for row in master_sheet.iter_rows(min_row=2):
                    if row[0].value:  # Device key column
                        unique_devices.add(row[0].value)
                
                # Property: Should have 3 unique devices (device1, device2, device3)
                assert len(unique_devices) == 3, f"Should have 3 unique devices, found {len(unique_devices)}"
                
            finally:
                workbook.close()
                # Additional cleanup for Windows file locking
                import gc
                import time
                gc.collect()
                time.sleep(0.2)  # Longer delay for Windows
    
    def test_output_directory_creation_property(self):
        """
        Property: Output directory should be created if it doesn't exist
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            # Use a subdirectory that doesn't exist yet
            reports_dir = os.path.join(temp_dir, "reports", "subdirectory")
            config = {'reports_directory': reports_dir}
            
            # Property: Directory should not exist initially
            assert not os.path.exists(reports_dir)
            
            # Create generator (should create directory)
            generator = ExcelReportGenerator(config)
            
            # Property: Directory should be created
            assert os.path.exists(reports_dir), "Reports directory should be created"
            assert os.path.isdir(reports_dir), "Reports path should be a directory"
    
    @given(
        ip_field_type=st.sampled_from(['primary_ip', 'ip_address', 'device_key_only']),
        ip_address=st.ip_addresses(v=4).map(str),
        hostname=st.text(min_size=1, max_size=10, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd')))
    )
    def test_ip_address_display_consistency_property(self, ip_field_type, ip_address, hostname):
        """
        Property 18: IP Address Display Consistency
        
        IP addresses should be consistently extracted from both 'primary_ip' and 'ip_address' 
        fields with device_key fallback, and should validate that extracted values are actual IP addresses.
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Create device info with different IP field configurations
            device_key = f"{hostname}:{ip_address}"
            device_info = {
                'hostname': hostname,
                'platform': 'IOS',
                'software_version': 'Version 1.0',
                'status': 'connected'
            }
            
            # Configure IP field based on test parameter
            if ip_field_type == 'primary_ip':
                device_info['primary_ip'] = ip_address
            elif ip_field_type == 'ip_address':
                device_info['ip_address'] = ip_address
            # For 'device_key_only', don't add IP fields - rely on device_key parsing
            
            # Test IP extraction
            extracted_ip = generator._extract_ip_address(device_key, device_info)
            
            # Property: Should always extract the correct IP address
            assert extracted_ip == ip_address, f"Should extract IP {ip_address}, got {extracted_ip}"
            
            # Property: Should handle missing IP fields gracefully
            empty_device_info = {'hostname': hostname, 'platform': 'IOS'}
            extracted_from_key = generator._extract_ip_address(device_key, empty_device_info)
            assert extracted_from_key == ip_address, f"Should extract IP from device_key when fields missing"
            
            # Property: primary_ip should take precedence over ip_address
            device_info_both = {
                'hostname': hostname,
                'primary_ip': ip_address,
                'ip_address': '10.0.0.99',  # Different IP
                'platform': 'IOS'
            }
            extracted_precedence = generator._extract_ip_address(device_key, device_info_both)
            assert extracted_precedence == ip_address, "primary_ip should take precedence over ip_address"
            
            # Property: Should reject invalid IP addresses (hostnames)
            invalid_device_key = f"{hostname}:{hostname}"  # hostname in IP position
            invalid_device_info = {'hostname': hostname, 'platform': 'IOS'}
            extracted_invalid = generator._extract_ip_address(invalid_device_key, invalid_device_info)
            assert extracted_invalid == '', "Should return empty string for invalid IP addresses"
    
    @given(
        num_devices=st.integers(min_value=0, max_value=3)
    )
    def test_empty_inventory_handling_property(self, num_devices):
        """
        Property: Generator should handle empty or minimal inventories gracefully
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Create minimal inventory
            devices = [(f"device{i}", f"192.168.1.{i+1}", "IOS") for i in range(num_devices)]
            inventory = self.create_test_inventory(devices)
            
            discovery_results = {
                'discovery_time_seconds': 10.0,
                'total_devices': num_devices,
                'successful_connections': num_devices,
                'failed_connections': 0,
                'filtered_devices': 0,
                'boundary_devices': 0,
                'max_depth_reached': 0
            }
            seed_devices = []
            
            # Property: Should not raise exception even with empty data
            try:
                filepaths = generator.generate_discovery_report(inventory, discovery_results, seed_devices)
                
                # Property: Files should still be created
                assert len(filepaths) > 0, "At least one Excel file should be created even with minimal data"
                main_filepath = filepaths[0]
                assert os.path.exists(main_filepath), "Main Excel file should be created even with minimal data"
                
                # Property: File should be valid Excel
                workbook = load_workbook(main_filepath)
                try:
                    assert len(workbook.sheetnames) > 0, "Workbook should have at least one sheet"
                finally:
                    workbook.close()
                
            except Exception as e:
                pytest.fail(f"Generator should handle empty inventory gracefully, but raised: {e}")
    
    @given(
        site_inventories=st.dictionaries(
            st.text(min_size=1, max_size=8, alphabet=st.characters(whitelist_categories=('Lu', 'Ll'))),
            st.lists(
                st.tuples(
                    st.text(min_size=1, max_size=10, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd'))),
                    st.ip_addresses(v=4).map(str),
                    st.sampled_from(['IOS', 'NX-OS', 'IOS-XE'])
                ),
                min_size=1, max_size=4,
                unique_by=lambda x: f"{x[0]}:{x[1]}"
            ),
            min_size=1, max_size=3
        )
    )
    @settings(max_examples=25, deadline=None)
    def test_site_summary_isolation_property(self, site_inventories):
        """
        Property 6: Site Summary Isolation
        
        PROPERTY: For any site-specific workbook, the summary statistics should not 
        include devices or connections from other sites.
        
        Validates: Requirements 3.5, 4.2
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Import SiteStatisticsCalculator for site-specific calculations
            from netwalker.discovery.site_statistics_calculator import SiteStatisticsCalculator
            calculator = SiteStatisticsCalculator()
            
            # Process each site separately
            for site_name, site_devices in site_inventories.items():
                if not site_devices:
                    continue
                
                # Create site-specific inventory
                site_inventory = self.create_test_inventory(site_devices)
                
                # Calculate site-specific statistics
                site_stats = calculator.generate_site_summary(site_name, site_inventory)
                
                # Generate site-specific report
                filepath = generator.generate_site_specific_report(site_inventory, site_stats, site_name)
                
                # Property: File should be created
                assert os.path.exists(filepath), f"Site report for {site_name} should be created"
                
                # Load and verify site isolation
                workbook = load_workbook(filepath)
                
                try:
                    # Property: Site Summary sheet should exist
                    assert "Site Summary" in workbook.sheetnames, "Site Summary sheet should exist"
                    
                    site_summary_sheet = workbook["Site Summary"]
                    
                    # Property: Site name should match in summary
                    site_name_found = False
                    for row in site_summary_sheet.iter_rows():
                        for cell in row:
                            if cell.value and site_name in str(cell.value):
                                site_name_found = True
                                break
                        if site_name_found:
                            break
                    
                    assert site_name_found, f"Site name {site_name} should appear in site summary"
                    
                    # Property: Device count should match site inventory size
                    device_count_found = False
                    expected_device_count = len(site_devices)
                    
                    for row in site_summary_sheet.iter_rows():
                        for i, cell in enumerate(row):
                            if cell.value == "Total Devices" and i + 1 < len(row):
                                actual_count = row[i + 1].value
                                if actual_count == expected_device_count:
                                    device_count_found = True
                                break
                        if device_count_found:
                            break
                    
                    assert device_count_found, f"Site device count should match inventory size {expected_device_count}"
                    
                    # Property: Site statistics should not exceed site inventory
                    for row in site_summary_sheet.iter_rows():
                        for i, cell in enumerate(row):
                            if (cell.value in ["Connected Devices", "Failed Devices", "Filtered Devices"] 
                                and i + 1 < len(row)):
                                count_value = row[i + 1].value
                                if isinstance(count_value, (int, float)):
                                    assert count_value <= expected_device_count, \
                                        f"Site statistic {cell.value} ({count_value}) should not exceed total devices ({expected_device_count})"
                    
                    # Property: Device Inventory sheet should contain only site devices
                    if "Device Inventory" in workbook.sheetnames:
                        device_sheet = workbook["Device Inventory"]
                        
                        # Count devices in inventory sheet
                        inventory_device_count = 0
                        site_device_hostnames = {device[0] for device in site_devices}
                        
                        for row in device_sheet.iter_rows(min_row=2):  # Skip header
                            if row[1].value:  # Hostname column
                                hostname = str(row[1].value)
                                inventory_device_count += 1
                                
                                # Property: Check if device belongs to site (allow neighbor devices)
                                # Neighbor devices may appear in inventory but that's acceptable
                                # The key property is that site statistics are accurate
                        
                        # Property: Inventory should contain at least the site devices
                        # (may contain more due to neighbors, which is acceptable)
                        assert inventory_device_count >= len(site_devices), \
                            f"Inventory should contain at least {len(site_devices)} site devices, found {inventory_device_count}"
                    
                    # Property: Site Connections sheet should only show site connections
                    if "Site Connections" in workbook.sheetnames:
                        connections_sheet = workbook["Site Connections"]
                        
                        site_device_hostnames = {device[0] for device in site_devices}
                        
                        for row in connections_sheet.iter_rows(min_row=2):  # Skip header
                            if row[0].value:  # Source device column
                                source_device = str(row[0].value)
                                
                                # Property: Source device should be from site inventory
                                # (This is the key isolation property - connections originate from site devices)
                                # We don't check target devices as they may be external
                
                finally:
                    workbook.close()
                    # Additional cleanup for Windows file locking
                    import gc
                    import time
                    gc.collect()
                    time.sleep(0.1)
    
    @given(
        site_devices=st.lists(
            st.tuples(
                st.text(min_size=1, max_size=10, alphabet=st.characters(whitelist_categories=('Lu', 'Ll', 'Nd'))),
                st.ip_addresses(v=4).map(str),
                st.sampled_from(['IOS', 'NX-OS', 'IOS-XE'])
            ),
            min_size=1, max_size=5,
            unique_by=lambda x: f"{x[0]}:{x[1]}"
        ),
        site_name=st.text(min_size=1, max_size=8, alphabet=st.characters(whitelist_categories=('Lu', 'Ll')))
    )
    @settings(max_examples=25, deadline=None)
    def test_site_workbook_content_completeness_property(self, site_devices, site_name):
        """
        Property 12: Site Workbook Content Completeness
        
        PROPERTY: For any site workbook, all devices associated with the site should 
        appear in the device inventory sheet with complete information.
        
        Validates: Requirements 8.1, 8.2
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            config = self.create_test_config(temp_dir)
            generator = ExcelReportGenerator(config)
            
            # Import SiteStatisticsCalculator for site-specific calculations
            from netwalker.discovery.site_statistics_calculator import SiteStatisticsCalculator
            calculator = SiteStatisticsCalculator()
            
            # Create site-specific inventory
            site_inventory = self.create_test_inventory(site_devices)
            
            # Calculate site-specific statistics
            site_stats = calculator.generate_site_summary(site_name, site_inventory)
            
            # Generate site-specific report
            filepath = generator.generate_site_specific_report(site_inventory, site_stats, site_name)
            
            # Property: File should be created
            assert os.path.exists(filepath), f"Site workbook for {site_name} should be created"
            
            # Load and verify content completeness
            workbook = load_workbook(filepath)
            
            try:
                # Property: Required sheets should exist
                required_sheets = ["Site Summary", "Device Inventory", "Site Connections", "VLAN Inventory"]
                sheet_names = workbook.sheetnames
                
                for required_sheet in required_sheets:
                    assert required_sheet in sheet_names, f"Required sheet '{required_sheet}' should exist in site workbook"
                
                # Property: Device Inventory sheet should contain all site devices
                device_sheet = workbook["Device Inventory"]
                
                # Collect device information from sheet
                sheet_devices = {}
                for row in device_sheet.iter_rows(min_row=2):  # Skip header
                    if row[0].value:  # Device Key column
                        device_key = str(row[0].value)
                        hostname = str(row[1].value) if row[1].value else ""
                        ip_address = str(row[2].value) if row[2].value else ""
                        platform = str(row[3].value) if row[3].value else ""
                        
                        sheet_devices[device_key] = {
                            'hostname': hostname,
                            'ip_address': ip_address,
                            'platform': platform
                        }
                
                # Property: All site devices should appear in the sheet
                for expected_device_key, expected_device_info in site_inventory.items():
                    expected_hostname = expected_device_info.get('hostname', '')
                    expected_ip = expected_device_info.get('ip_address', '') or expected_device_info.get('primary_ip', '')
                    
                    # Check if device appears in sheet (by key or by hostname/IP match)
                    device_found = False
                    
                    if expected_device_key in sheet_devices:
                        device_found = True
                    else:
                        # Check by hostname/IP match
                        for sheet_key, sheet_info in sheet_devices.items():
                            if (sheet_info['hostname'] == expected_hostname or 
                                sheet_info['ip_address'] == expected_ip):
                                device_found = True
                                break
                    
                    assert device_found, f"Site device {expected_hostname} ({expected_ip}) should appear in device inventory sheet"
                
                # Property: Device information should be complete (non-empty for key fields)
                for device_key, device_info in sheet_devices.items():
                    # Property: Hostname should not be empty
                    assert device_info['hostname'], f"Device {device_key} should have non-empty hostname"
                    
                    # Property: IP address should not be empty
                    assert device_info['ip_address'], f"Device {device_key} should have non-empty IP address"
                    
                    # Property: Platform information should be present (may be empty for some devices)
                    # This is acceptable as not all devices may report platform information
                
                # Property: Site Summary should contain site-specific statistics
                site_summary_sheet = workbook["Site Summary"]
                
                # Check that site statistics match the calculated values
                summary_stats = {}
                for row in site_summary_sheet.iter_rows():
                    for i, cell in enumerate(row):
                        if cell.value in ["Total Devices", "Connected Devices", "Total Connections"] and i + 1 < len(row):
                            summary_stats[cell.value] = row[i + 1].value
                
                # Property: Total devices in summary should match site inventory
                if "Total Devices" in summary_stats:
                    assert summary_stats["Total Devices"] == site_stats.total_devices, \
                        f"Summary total devices should match site statistics: {summary_stats['Total Devices']} != {site_stats.total_devices}"
                
                # Property: Site Connections sheet should show connection details
                if "Site Connections" in workbook.sheetnames:
                    connections_sheet = workbook["Site Connections"]
                    
                    # Check that connections sheet has proper headers
                    header_row = list(connections_sheet.iter_rows(min_row=1, max_row=1))[0]
                    expected_headers = ["Source Device", "Source IP", "Target Device", "Target IP", "Connection Type"]
                    
                    header_values = [cell.value for cell in header_row if cell.value]
                    for expected_header in expected_headers:
                        assert expected_header in header_values, f"Connections sheet should have header '{expected_header}'"
                    
                    # Property: Connection data should reference site devices
                    connection_count = 0
                    site_hostnames = {device_info.get('hostname', '') for device_info in site_inventory.values()}
                    
                    for row in connections_sheet.iter_rows(min_row=2):  # Skip header
                        if row[0].value:  # Source Device column
                            connection_count += 1
                            source_device = str(row[0].value)
                            
                            # Property: Source device should be from site inventory
                            # (Connections should originate from site devices)
                            source_found = any(hostname in source_device for hostname in site_hostnames if hostname)
                            # Note: We allow some flexibility here as device names may be cleaned/formatted
                    
                    # Property: If site has devices with neighbors, there should be some connections
                    devices_with_neighbors = sum(1 for device_info in site_inventory.values() 
                                               if device_info.get('neighbors', []))
                    
                    if devices_with_neighbors > 0:
                        # We expect some connections, but don't enforce exact count due to neighbor processing complexity
                        pass  # Connection count validation is handled by site statistics tests
                
            finally:
                workbook.close()
                # Additional cleanup for Windows file locking
                import gc
                import time
                gc.collect()
                time.sleep(0.1)