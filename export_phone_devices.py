"""Export all devices with phone capability to a spreadsheet."""
import pyodbc
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

server = "eit-prisqldb01.tgna.tegna.com"
database = "NetWalker"
username = "NetWalker"
password = base64.b64decode("Rmx1ZmZ5QnVubnlIaXRieWFCdXM=").decode()

conn_str = (
    f"DRIVER={{SQL Server}};"
    f"SERVER={server};"
    f"DATABASE={database};"
    f"UID={username};"
    f"PWD={password};"
    f"TrustServerCertificate=yes"
)

conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

cursor.execute("""
    SELECT device_name, platform, hardware_model, capabilities, 
           first_seen, last_seen, status
    FROM devices
    WHERE capabilities LIKE '%phone%'
       OR device_name LIKE 'SEP%'
       OR platform LIKE '%phone%'
    ORDER BY device_name
""")

rows = cursor.fetchall()
print(f"Found {len(rows)} phone devices")

wb = Workbook()
ws = wb.active
ws.title = "Phone Devices"

headers = ["Device Name", "Platform", "Hardware Model", "Capabilities", 
           "First Seen", "Last Seen", "Status"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row in enumerate(rows, 2):
    ws.cell(row=row_idx, column=1, value=row.device_name)
    ws.cell(row=row_idx, column=2, value=row.platform or "")
    ws.cell(row=row_idx, column=3, value=row.hardware_model or "")
    ws.cell(row=row_idx, column=4, value=row.capabilities or "")
    ws.cell(row=row_idx, column=5, value=str(row.first_seen) if row.first_seen else "")
    ws.cell(row=row_idx, column=6, value=str(row.last_seen) if row.last_seen else "")
    ws.cell(row=row_idx, column=7, value=row.status or "")

for col in range(1, len(headers) + 1):
    ws.column_dimensions[chr(64 + col)].width = 25

timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
filename = f"Phone_Devices_{timestamp}.xlsx"
wb.save(filename)
print(f"Exported to {filename}")

cursor.close()
conn.close()
