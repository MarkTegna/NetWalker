#!/usr/bin/env python3
"""
Check if WTSP-CORE-A appears in the BORO-CORE-A seed report.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def check_seed_report():
    """Check the seed report for WTSP-CORE-A"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("CHECKING BORO-CORE-A SEED REPORT FOR WTSP-CORE-A")
    logger.info("=" * 80)
    
    try:
        from openpyxl import load_workbook
        
        # Find the most recent seed report
        reports_dir = Path("dist/NetWalker_v0.3.29/reports")
        seed_reports = list(reports_dir.glob("Seed_BORO-CORE-A_*.xlsx"))
        
        if not seed_reports:
            logger.error("No BORO-CORE-A seed reports found")
            return False
        
        latest_seed = max(seed_reports, key=lambda p: p.stat().st_mtime)
        logger.info(f"Analyzing: {latest_seed.name}\n")
        
        workbook = load_workbook(latest_seed)
        
        # Check the Neighbors Overview sheet
        if "Neighbors Overview" in workbook.sheetnames:
            ws = workbook["Neighbors Overview"]
            logger.info("Checking 'Neighbors Overview' sheet...")
            
            wtsp_found = False
            neighbors = []
            
            for row in ws.iter_rows(min_row=6, values_only=True):  # Skip headers
                if row and row[0]:  # If first column has data
                    hostname = str(row[0]) if row[0] else ''
                    neighbors.append(hostname)
                    if 'WTSP-CORE-A' in hostname.upper():
                        wtsp_found = True
                        logger.info(f"  ✅ Found WTSP-CORE-A in neighbors: {hostname}")
                        logger.info(f"     IP: {row[1] if len(row) > 1 else 'N/A'}")
                        logger.info(f"     Protocol: {row[2] if len(row) > 2 else 'N/A'}")
                        logger.info(f"     Platform: {row[5] if len(row) > 5 else 'N/A'}")
            
            if not wtsp_found:
                logger.error("  ❌ WTSP-CORE-A NOT found in neighbors list")
                logger.info(f"\n  Total neighbors found: {len(neighbors)}")
                logger.info(f"  First 10 neighbors:")
                for i, neighbor in enumerate(neighbors[:10], 1):
                    logger.info(f"    {i}. {neighbor}")
        
        # Check if there's a WTSP-CORE-A neighbor detail sheet
        wtsp_sheet_found = False
        for sheet_name in workbook.sheetnames:
            if 'WTSP' in sheet_name.upper():
                wtsp_sheet_found = True
                logger.info(f"\n  ✅ Found WTSP sheet: {sheet_name}")
        
        if not wtsp_sheet_found:
            logger.info(f"\n  ❌ No WTSP neighbor detail sheet found")
        
        workbook.close()
        
        logger.info(f"\n{'=' * 80}")
        logger.info("CONCLUSION")
        logger.info(f"{'=' * 80}")
        
        if wtsp_found:
            logger.info("WTSP-CORE-A is listed as a neighbor of BORO-CORE-A")
            logger.info("But it's not in the main inventory, which means:")
            logger.info("  1. It was discovered as a neighbor")
            logger.info("  2. But it was NOT walked (neighbors not discovered)")
            logger.info("  3. Likely due to depth limit or filtering")
        else:
            logger.error("WTSP-CORE-A is NOT even listed as a neighbor")
            logger.error("This means it's not being discovered at all from BORO-CORE-A")
        
        return True
        
    except Exception as e:
        logger.error(f"Error checking seed report: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = check_seed_report()
    sys.exit(0 if success else 1)