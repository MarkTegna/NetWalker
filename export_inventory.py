#!/usr/bin/env python3
"""
Export network device inventory from database to Excel spreadsheet
"""

import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Add the project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from netwalker.config.config_manager import ConfigurationManager
from netwalker.database import DatabaseManager


def main():
    """Main function"""
    
    # Load configuration
    config_file = 'netwalker.ini'
    config_manager = ConfigurationManager(config_file)
    parsed_config = config_manager.load_configuration()
    db_config = parsed_config.get('database', {})
    
    # Initialize database manager
    db_manager = DatabaseManager(db_config)
    
    if not db_manager.connect():
        print("[FAIL] Could not connect to database")
        return 1
    
    try:
        cursor = db_manager.connection.cursor()
        
        print("\n" + "=" * 80)
        print("Exporting Network Device Inventory")
        print("=" * 80)
        
        # Query all devices with their primary IP
        query = """
            SELECT 
                d.device_name,
                d.platform,
                d.capabilities,
                d.hardware_model,
                d.serial_number,
                (SELECT TOP 1 ip_address
                 FROM device_interfaces
                 WHERE device_id = d.device_id
                 ORDER BY
                    CASE
                        WHEN interface_name LIKE '%Management%' THEN 1
                        WHEN interface_name LIKE '%Loopback%' THEN 2
                        WHEN interface_name LIKE '%Vlan%' THEN 3
                        ELSE 4
                    END,
                    interface_name) as primary_ip,
                d.status,
                d.first_seen,
                d.last_seen
            FROM devices d
            ORDER BY d.device_name
        """
        
        cursor.execute(query)
        devices = cursor.fetchall()
        
        print(f"\nRetrieved {len(devices)} devices from database")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Device Inventory"
        
        # Define headers
        headers = [
            "Device Name",
            "Primary IP",
            "Platform",
            "Capabilities",
            "Hardware Model",
            "Serial Number",
            "Status",
            "First Seen",
            "Last Seen"
        ]
        
        # Write headers with formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write device data
        for row_num, device in enumerate(devices, 2):
            device_name = device[0] or ''
            platform = device[1] or 'Unknown'
            capabilities = device[2] or ''
            hardware_model = device[3] or ''
            serial_number = device[4] or ''
            primary_ip = device[5] or ''
            status = device[6] or ''
            first_seen = device[7]
            last_seen = device[8]
            
            # Format timestamps (they may already be strings or datetime objects)
            if first_seen:
                if hasattr(first_seen, 'strftime'):
                    first_seen_str = first_seen.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    first_seen_str = str(first_seen)
            else:
                first_seen_str = ''
            
            if last_seen:
                if hasattr(last_seen, 'strftime'):
                    last_seen_str = last_seen.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    last_seen_str = str(last_seen)
            else:
                last_seen_str = ''
            
            # Write row data
            ws.cell(row=row_num, column=1).value = device_name
            ws.cell(row=row_num, column=2).value = primary_ip
            ws.cell(row=row_num, column=3).value = platform
            ws.cell(row=row_num, column=4).value = capabilities
            ws.cell(row=row_num, column=5).value = hardware_model
            ws.cell(row=row_num, column=6).value = serial_number
            ws.cell(row=row_num, column=7).value = status
            ws.cell(row=row_num, column=8).value = first_seen_str
            ws.cell(row=row_num, column=9).value = last_seen_str
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d-%H-%M')
        filename = f"Network_Inventory_{timestamp}.xlsx"
        
        # Save workbook
        wb.save(filename)
        
        print(f"\n[OK] Inventory exported successfully")
        print(f"[OK] File: {filename}")
        print(f"[OK] Total devices: {len(devices)}")
        print("\n" + "=" * 80)
        
        cursor.close()
        
    except Exception as e:
        print(f"\n[FAIL] Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    finally:
        db_manager.disconnect()
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
