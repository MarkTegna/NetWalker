#!/usr/bin/env python3
"""
Export all devices with connection failures > 0 to Excel spreadsheet.
"""

import pyodbc
import configparser
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Read database configuration from netwalker.ini
config = configparser.ConfigParser()
config.read('netwalker.ini')

# Decrypt password (simple base64 encoding)
encrypted_password = config.get('database', 'password').replace('ENC:', '')
password = base64.b64decode(encrypted_password).decode('utf-8')

# Database configuration
DB_SERVER = config.get('database', 'server')
DB_NAME = config.get('database', 'database')
DB_USER = config.get('database', 'username')
DB_PASSWORD = password
DB_PORT = config.get('database', 'port')

def export_failures():
    """Export devices with connection failures to Excel."""
    # Build connection string
    conn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={DB_SERVER},{DB_PORT};"
        f"DATABASE={DB_NAME};"
        f"UID={DB_USER};"
        f"PWD={DB_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    
    try:
        # Query devices with connection failures > 0
        cursor.execute("""
            SELECT 
                d.device_id,
                d.device_name,
                (SELECT TOP 1 ip_address FROM device_interfaces WHERE device_id = d.device_id) as ip_address,
                d.serial_number,
                d.platform,
                d.hardware_model,
                d.capabilities,
                d.connection_failures,
                d.first_seen,
                d.last_seen,
                d.status
            FROM devices d
            WHERE d.connection_failures > 0
            ORDER BY d.connection_failures DESC, d.device_name
        """)
        
        devices = cursor.fetchall()
        
        if not devices:
            print("No devices with connection failures found")
            return
        
        print(f"\nFound {len(devices)} devices with connection failures")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Connection Failures"
        
        # Define headers
        headers = [
            'Device ID', 'Device Name', 'IP Address', 'Serial Number', 
            'Platform', 'Hardware Model', 'Capabilities', 'Connection Failures',
            'First Seen', 'Last Seen', 'Status'
        ]
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, device in enumerate(devices, start=2):
            ws.cell(row=row_idx, column=1, value=device.device_id)
            ws.cell(row=row_idx, column=2, value=device.device_name)
            ws.cell(row=row_idx, column=3, value=device.ip_address or 'N/A')
            ws.cell(row=row_idx, column=4, value=device.serial_number)
            ws.cell(row=row_idx, column=5, value=device.platform)
            ws.cell(row=row_idx, column=6, value=device.hardware_model)
            ws.cell(row=row_idx, column=7, value=device.capabilities)
            ws.cell(row=row_idx, column=8, value=device.connection_failures)
            ws.cell(row=row_idx, column=9, value=str(device.first_seen) if device.first_seen else '')
            ws.cell(row=row_idx, column=10, value=str(device.last_seen) if device.last_seen else '')
            ws.cell(row=row_idx, column=11, value=device.status)
            
            # Highlight rows with 5+ failures in red
            if device.connection_failures >= 5:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col).column_letter
            
            for row in range(1, len(devices) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d-%H-%M')
        filename = f'Connection_Failures_{timestamp}.xlsx'
        
        # Save workbook
        wb.save(filename)
        
        print(f"\n{'='*70}")
        print(f"EXPORT COMPLETE")
        print(f"{'='*70}")
        print(f"File: {filename}")
        print(f"Total devices: {len(devices)}")
        print(f"Devices with 5+ failures: {sum(1 for d in devices if d.connection_failures >= 5)}")
        print(f"\nTimestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
    except Exception as e:
        print(f"\n[ERROR] Failed to export: {e}")
        raise
    
    finally:
        conn.close()

if __name__ == '__main__':
    export_failures()
