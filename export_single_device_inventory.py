#!/usr/bin/env python3
"""
Export Single Device Inventory from NetWalker Database

Queries the database and creates an Excel spreadsheet for a specific device.

Author: Mark Oldham
"""

import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Add the project root to Python path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from netwalker.config.config_manager import ConfigurationManager
from netwalker.database import DatabaseManager


def create_device_inventory_sheet(wb, db_manager, device_name):
    """Create device inventory sheet"""
    
    ws = wb.active
    ws.title = "Device Inventory"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers
    headers = [
        "Device Name",
        "Platform",
        "Hardware Model",
        "Serial Number",
        "Software Version",
        "IP Address",
        "Capabilities",
        "Stack Members",
        "First Seen",
        "Last Seen",
        "Status"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12
    
    # Query device
    cursor = db_manager.connection.cursor()
    
    query = """
        SELECT
            d.device_name,
            d.platform,
            d.hardware_model,
            d.serial_number,
            d.capabilities,
            d.first_seen,
            d.last_seen,
            d.status,
            (SELECT TOP 1 software_version 
             FROM device_versions 
             WHERE device_id = d.device_id 
             ORDER BY last_seen DESC) as software_version,
            (SELECT TOP 1 ip_address
             FROM device_interfaces
             WHERE device_id = d.device_id
             ORDER BY
                CASE
                    WHEN interface_name LIKE '%Management%' THEN 1
                    WHEN interface_name LIKE '%Loopback%' THEN 2
                    WHEN interface_name LIKE '%Vlan%' THEN 3
                    ELSE 4
                END,
                interface_name) as ip_address,
            (SELECT COUNT(*)
             FROM device_stack_members
             WHERE device_id = d.device_id) as stack_member_count
        FROM devices d
        WHERE d.status = 'active'
          AND d.device_name = ?
    """
    
    cursor.execute(query, (device_name,))
    device = cursor.fetchone()
    
    if device:
        device_name_val = device[0]
        platform = device[1] or ''
        hardware_model = device[2] or ''
        serial_number = device[3] or ''
        capabilities = device[4] or ''
        first_seen = device[5]
        last_seen = device[6]
        status = device[7]
        software_version = device[8] or ''
        ip_address = device[9] or ''
        stack_member_count = device[10] or 0
        
        # Format stack member count
        stack_display = str(stack_member_count) if stack_member_count > 0 else ''
        
        # Format dates
        first_seen_str = str(first_seen) if first_seen else ''
        last_seen_str = str(last_seen) if last_seen else ''
        
        # Write row
        row_data = [
            device_name_val,
            platform,
            hardware_model,
            serial_number,
            software_version,
            ip_address,
            capabilities,
            stack_display,
            first_seen_str,
            last_seen_str,
            status
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = border
    
    cursor.close()
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    return 1 if device else 0


def create_stack_members_sheet(wb, db_manager, device_name):
    """Create stack members sheet"""
    
    ws = wb.create_sheet("Stack Members")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers
    headers = [
        "Stack Device",
        "Stack IP",
        "Platform",
        "Switch Number",
        "Role",
        "Priority",
        "Hardware Model",
        "Serial Number",
        "MAC Address",
        "Software Version",
        "State",
        "First Seen",
        "Last Seen"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    
    # Query stack members - get only the most recent record for each switch
    cursor = db_manager.connection.cursor()
    
    query = """
        SELECT
            d.device_name,
            d.platform,
            sm.switch_number,
            sm.role,
            sm.priority,
            sm.hardware_model,
            sm.serial_number,
            sm.mac_address,
            sm.software_version,
            sm.state,
            sm.first_seen,
            sm.last_seen,
            (SELECT TOP 1 ip_address
             FROM device_interfaces
             WHERE device_id = d.device_id
             ORDER BY
                CASE
                    WHEN interface_name LIKE '%Management%' THEN 1
                    WHEN interface_name LIKE '%Loopback%' THEN 2
                    WHEN interface_name LIKE '%Vlan%' THEN 3
                    ELSE 4
                END,
                interface_name) as ip_address
        FROM device_stack_members sm
        INNER JOIN devices d ON sm.device_id = d.device_id
        WHERE d.status = 'active'
          AND d.device_name = ?
          AND sm.stack_member_id IN (
              SELECT MAX(stack_member_id)
              FROM device_stack_members sm2
              WHERE sm2.device_id = sm.device_id
                AND sm2.switch_number = sm.switch_number
              GROUP BY device_id, switch_number
          )
        ORDER BY sm.switch_number
    """
    
    cursor.execute(query, (device_name,))
    members = cursor.fetchall()
    
    # Write stack member data
    row_num = 2
    for member in members:
        device_name_val = member[0]
        platform = member[1] or ''
        switch_number = member[2]
        role = member[3] or ''
        priority = member[4] if member[4] is not None else ''
        hardware_model = member[5] or ''
        serial_number = member[6] or ''
        mac_address = member[7] or ''
        software_version = member[8] or ''
        state = member[9] or ''
        first_seen = member[10]
        last_seen = member[11]
        ip_address = member[12] or ''
        
        # Format dates
        first_seen_str = str(first_seen) if first_seen else ''
        last_seen_str = str(last_seen) if last_seen else ''
        
        # Write row
        row_data = [
            device_name_val,
            ip_address,
            platform,
            switch_number,
            role,
            priority,
            hardware_model,
            serial_number,
            mac_address,
            software_version,
            state,
            first_seen_str,
            last_seen_str
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = border
        
        row_num += 1
    
    cursor.close()
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    return len(members)


def main():
    """Main function"""
    
    if len(sys.argv) < 2:
        print("Usage: python export_single_device_inventory.py <device_name>")
        print("Example: python export_single_device_inventory.py KXTV-MC-STACK1")
        return 1
    
    device_name = sys.argv[1]
    
    # Load configuration
    config_file = 'netwalker.ini'
    config_manager = ConfigurationManager(config_file)
    parsed_config = config_manager.load_configuration()
    db_config = parsed_config.get('database', {})
    
    # Initialize database manager
    db_manager = DatabaseManager(db_config)
    
    if not db_manager.connect():
        print("[FAIL] Could not connect to database")
        return 1
    
    try:
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d-%H%M')
        output_file = f"{device_name}_Inventory_{timestamp}.xlsx"
        
        print(f"\nGenerating inventory spreadsheet for {device_name}...")
        print(f"Output file: {output_file}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create device inventory sheet
        device_count = create_device_inventory_sheet(wb, db_manager, device_name)
        
        if device_count == 0:
            print(f"\n[FAIL] Device '{device_name}' not found in database")
            return 1
        
        # Create stack members sheet
        stack_count = create_stack_members_sheet(wb, db_manager, device_name)
        
        # Save workbook
        wb.save(output_file)
        
        print(f"\n[OK] Inventory exported successfully")
        print(f"  Device: {device_name}")
        print(f"  Stack members: {stack_count}")
        print(f"  File: {output_file}")
        
    except Exception as e:
        print(f"\n[FAIL] Error exporting inventory: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    finally:
        db_manager.disconnect()
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
