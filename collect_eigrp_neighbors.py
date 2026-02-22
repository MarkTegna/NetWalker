"""
Connect to all -UW devices and collect EIGRP neighbor information
Author: Mark Oldham
"""

import pyodbc
import base64
from scrapli.driver.core import IOSXEDriver
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import configparser

def decode_password(encoded_password):
    """Decode the encrypted password"""
    if encoded_password.startswith("ENC:"):
        encoded_password = encoded_password[4:]
    return base64.b64decode(encoded_password).decode('utf-8')

def connect_to_database():
    """Connect to the NetWalker database"""
    server = "eit-prisqldb01.tgna.tegna.com"
    port = 1433
    database = "NetWalker"
    username = "NetWalker"
    encoded_password = "ENC:Rmx1ZmZ5QnVubnlIaXRieWFCdXM="
    
    password = decode_password(encoded_password)
    
    conn_str = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={server},{port};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        f"TrustServerCertificate=yes;"
    )
    
    try:
        connection = pyodbc.connect(conn_str, timeout=30)
        return connection
    except Exception as e:
        print(f"Error connecting to database: {e}")
        return None

def get_uw_devices():
    """Get all devices with -UW in their name"""
    conn = connect_to_database()
    if not conn:
        return []
    
    try:
        cursor = conn.cursor()
        
        # Get devices with -UW in name and their primary IP
        query = """
        SELECT DISTINCT
            d.device_name,
            di.ip_address
        FROM devices d
        LEFT JOIN (
            SELECT device_id, ip_address,
                   ROW_NUMBER() OVER (PARTITION BY device_id ORDER BY last_seen DESC) as rn
            FROM device_interfaces
        ) di ON d.device_id = di.device_id AND di.rn = 1
        WHERE d.device_name LIKE '%-%UW%'
        AND di.ip_address IS NOT NULL
        ORDER BY d.device_name
        """
        
        cursor.execute(query)
        devices = cursor.fetchall()
        
        return [(name, ip) for name, ip in devices]
        
    except Exception as e:
        print(f"Error getting devices: {e}")
        return []
    finally:
        conn.close()

def load_credentials():
    """Load credentials from netwalker.ini"""
    config = configparser.ConfigParser()
    config.read('netwalker.ini')
    
    # Try to get credentials from config
    username = config.get('credentials', 'username', fallback='admin')
    password = config.get('credentials', 'password', fallback='')
    enable_password = config.get('credentials', 'enable_password', fallback='')
    
    return username, password, enable_password

def connect_and_run_command(device_name, ip_address, username, password, enable_password):
    """Connect to device and run EIGRP command"""
    
    print(f"  Connecting to {device_name} ({ip_address})...")
    
    device = {
        "host": ip_address,
        "auth_username": username,
        "auth_password": password,
        "auth_secondary": enable_password,
        "auth_strict_key": False,
        "timeout_socket": 30,
        "timeout_transport": 30,
    }
    
    try:
        conn = IOSXEDriver(**device)
        conn.open()
        
        # Run the command
        response = conn.send_command("show ip eigrp vrf WAN neigh")
        
        conn.close()
        
        return {
            'device': device_name,
            'ip': ip_address,
            'status': 'Success',
            'output': response.result
        }
        
    except Exception as e:
        return {
            'device': device_name,
            'ip': ip_address,
            'status': 'Failed',
            'output': str(e)
        }

def parse_eigrp_output(output):
    """Parse EIGRP neighbor output"""
    neighbors = []
    
    if not output or 'Invalid' in output or 'not exist' in output:
        return neighbors
    
    lines = output.split('\n')
    
    for line in lines:
        line = line.strip()
        # Skip header lines and empty lines
        if not line or line.startswith('EIGRP') or line.startswith('H ') or line.startswith('---'):
            continue
        
        # Parse neighbor line
        parts = line.split()
        if len(parts) >= 5:
            try:
                h = parts[0]
                address = parts[1]
                interface = parts[2]
                hold = parts[3]
                uptime = parts[4]
                srtt = parts[5] if len(parts) > 5 else ''
                rto = parts[6] if len(parts) > 6 else ''
                q_cnt = parts[7] if len(parts) > 7 else ''
                seq_num = parts[8] if len(parts) > 8 else ''
                
                neighbors.append({
                    'H': h,
                    'Address': address,
                    'Interface': interface,
                    'Hold': hold,
                    'Uptime': uptime,
                    'SRTT': srtt,
                    'RTO': rto,
                    'Q_Cnt': q_cnt,
                    'Seq_Num': seq_num
                })
            except:
                pass
    
    return neighbors

def export_to_excel(results):
    """Export results to Excel"""
    
    timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
    output_file = f"EIGRP_Neighbors_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "EIGRP Neighbors"
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    headers = ['Device', 'Device IP', 'Status', 'Neighbor H', 'Neighbor Address', 
               'Interface', 'Hold', 'Uptime', 'SRTT', 'RTO', 'Q Cnt', 'Seq Num']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write data
    row_idx = 2
    for result in results:
        device = result['device']
        ip = result['ip']
        status = result['status']
        
        if status == 'Success':
            neighbors = parse_eigrp_output(result['output'])
            
            if neighbors:
                for neighbor in neighbors:
                    ws.cell(row=row_idx, column=1, value=device)
                    ws.cell(row=row_idx, column=2, value=ip)
                    ws.cell(row=row_idx, column=3, value=status)
                    ws.cell(row=row_idx, column=4, value=neighbor.get('H', ''))
                    ws.cell(row=row_idx, column=5, value=neighbor.get('Address', ''))
                    ws.cell(row=row_idx, column=6, value=neighbor.get('Interface', ''))
                    ws.cell(row=row_idx, column=7, value=neighbor.get('Hold', ''))
                    ws.cell(row=row_idx, column=8, value=neighbor.get('Uptime', ''))
                    ws.cell(row=row_idx, column=9, value=neighbor.get('SRTT', ''))
                    ws.cell(row=row_idx, column=10, value=neighbor.get('RTO', ''))
                    ws.cell(row=row_idx, column=11, value=neighbor.get('Q_Cnt', ''))
                    ws.cell(row=row_idx, column=12, value=neighbor.get('Seq_Num', ''))
                    row_idx += 1
            else:
                # No neighbors found
                ws.cell(row=row_idx, column=1, value=device)
                ws.cell(row=row_idx, column=2, value=ip)
                ws.cell(row=row_idx, column=3, value='No Neighbors')
                row_idx += 1
        else:
            # Failed connection
            ws.cell(row=row_idx, column=1, value=device)
            ws.cell(row=row_idx, column=2, value=ip)
            ws.cell(row=row_idx, column=3, value=status)
            ws.cell(row=row_idx, column=4, value=result['output'])
            row_idx += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"\nResults exported to: {output_file}")
    return output_file

def main():
    """Main function"""
    
    print("EIGRP Neighbor Collection for -UW Devices")
    print("=" * 100)
    
    # Get devices
    print("\nGetting devices with -UW in name from database...")
    devices = get_uw_devices()
    
    if not devices:
        print("No devices found with -UW in their name")
        return
    
    print(f"Found {len(devices)} devices")
    
    # Load credentials
    username, password, enable_password = load_credentials()
    
    if not password:
        print("\nCredentials not found in netwalker.ini")
        username = input("Enter username: ")
        password = input("Enter password: ")
        enable_password = input("Enter enable password: ")
    
    # Collect data
    print(f"\nConnecting to devices and collecting EIGRP neighbors...")
    print("=" * 100)
    
    results = []
    success_count = 0
    failed_count = 0
    
    for device_name, ip_address in devices:
        result = connect_and_run_command(device_name, ip_address, username, password, enable_password)
        results.append(result)
        
        if result['status'] == 'Success':
            success_count += 1
            neighbors = parse_eigrp_output(result['output'])
            print(f"    ✓ {device_name}: {len(neighbors)} neighbors found")
        else:
            failed_count += 1
            print(f"    ✗ {device_name}: {result['status']}")
    
    # Export to Excel
    print("\n" + "=" * 100)
    print(f"\nCollection Summary:")
    print(f"  Total devices: {len(devices)}")
    print(f"  Successful: {success_count}")
    print(f"  Failed: {failed_count}")
    
    export_to_excel(results)

if __name__ == "__main__":
    main()
