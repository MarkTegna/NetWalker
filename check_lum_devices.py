#!/usr/bin/env python3
"""
Check if LUM* devices were walked and if they have WTSP-CORE-A as a neighbor.
"""

import sys
import logging
from pathlib import Path

def setup_logging():
    """Setup logging for the test"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

def check_lum_seed_reports():
    """Check if LUM* devices have seed reports (meaning they were walked)"""
    logger = setup_logging()
    
    logger.info("=" * 80)
    logger.info("CHECKING LUM* DEVICE DISCOVERY")
    logger.info("=" * 80)
    
    try:
        from openpyxl import load_workbook
        
        reports_dir = Path("dist/NetWalker_v0.3.29/reports")
        
        # Look for LUM* seed reports
        lum_reports = list(reports_dir.glob("Seed_LUM*.xlsx"))
        
        logger.info(f"\nLUM* seed reports found: {len(lum_reports)}")
        
        if not lum_reports:
            logger.error("❌ No LUM* seed reports found!")
            logger.error("This means LUM* devices were NOT walked")
            logger.error("\nPossible reasons:")
            logger.error("  1. LUM* devices are at depth 2, but max_depth might be too low")
            logger.error("  2. LUM* devices are being filtered")
            logger.error("  3. LUM* devices failed to connect")
            
            # Check if they're in the inventory
            inventory_files = list(reports_dir.glob("Inventory_*.xlsx"))
            if inventory_files:
                latest_inventory = max(inventory_files, key=lambda p: p.stat().st_mtime)
                workbook = load_workbook(latest_inventory)
                
                for sheet_name in workbook.sheetnames:
                    if 'inventory' in sheet_name.lower():
                        ws = workbook[sheet_name]
                        
                        logger.info(f"\nChecking inventory for LUM* devices...")
                        lum_devices = []
                        
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and row[1]:  # Hostname column
                                hostname = str(row[1])
                                if 'LUM' in hostname.upper():
                                    status = row[8] if len(row) > 8 else 'unknown'
                                    lum_devices.append({
                                        'hostname': hostname,
                                        'ip': row[2] if len(row) > 2 else '',
                                        'status': status
                                    })
                        
                        if lum_devices:
                            logger.info(f"Found {len(lum_devices)} LUM* devices in inventory:")
                            for device in lum_devices:
                                logger.info(f"  - {device['hostname']} ({device['ip']}) - Status: {device['status']}")
                        else:
                            logger.error("  No LUM* devices found in inventory either!")
                        
                        break
                
                workbook.close()
            
            return False
        
        # Check each LUM* seed report for WTSP-CORE-A
        for lum_report in lum_reports:
            logger.info(f"\n{'-' * 60}")
            logger.info(f"Analyzing: {lum_report.name}")
            logger.info(f"{'-' * 60}")
            
            workbook = load_workbook(lum_report)
            
            # Check Neighbors Overview sheet
            if "Neighbors Overview" in workbook.sheetnames:
                ws = workbook["Neighbors Overview"]
                
                wtsp_found = False
                neighbors = []
                
                for row in ws.iter_rows(min_row=6, values_only=True):
                    if row and row[0]:
                        hostname = str(row[0]) if row[0] else ''
                        neighbors.append(hostname)
                        if 'WTSP' in hostname.upper():
                            wtsp_found = True
                            logger.info(f"  ✅ Found WTSP in neighbors: {hostname}")
                            logger.info(f"     IP: {row[1] if len(row) > 1 else 'N/A'}")
                            logger.info(f"     Protocol: {row[2] if len(row) > 2 else 'N/A'}")
                
                if not wtsp_found:
                    logger.info(f"  ❌ WTSP-CORE-A not found in {lum_report.name}")
                    logger.info(f"     Total neighbors: {len(neighbors)}")
                    if neighbors:
                        logger.info(f"     First 5 neighbors: {neighbors[:5]}")
            
            workbook.close()
        
        return True
        
    except Exception as e:
        logger.error(f"Error checking LUM reports: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = check_lum_seed_reports()
    sys.exit(0 if success else 1)